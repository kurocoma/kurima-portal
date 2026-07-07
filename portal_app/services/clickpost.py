from __future__ import annotations

import asyncio
import csv
import hashlib
import json
import os
import re
import unicodedata
import warnings as warning_module
from dataclasses import asdict, dataclass, replace
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable

import openpyxl

from portal_app.services.master_cache import cached_by_mtime
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright

from portal_app.services.next_engine_downloader import (
    APP_ROOT,
    STORAGE_STATE_PATH,
    NextEngineOrderDetailDownloader,
    _auto_accept_dialogs,
    _chromium_launch_options,
    _headless_default,
    _next_engine_storage_lock,
)
from portal_app.services.next_engine_order_status import (
    OrderStatusBatchRestoreResult,
    restore_next_engine_print_wait_batch_sync,
)
from portal_app.services.paths import PortalPaths, find_portal_paths
from portal_app.settings import download_timeout_ms, nav_timeout_ms


ORDER_LIST_PRINT_WAIT_URL = "https://main.next-engine.com/Userjyuchu/index?search_condi=17"
ORDER_DETAIL_LIST_URL = "https://main.next-engine.com/Userjyuchumeisai"
CLICKPOST_URL = "https://clickpost.jp/"
CLICKPOST_MYPAGE_URL = "https://clickpost.jp/mypage/index"
CLICKPOST_MULTIPLE_PRINT_URL = "https://clickpost.jp/labels/multiple_print"
YAHOO_LOGOUT_URL = "https://login.yahoo.co.jp/config/login?.done=https%3A%2F%2Fwww.yahoo.co.jp&.src=help&logout=1"
INVOICE_ACTION_ID = "#extension_execute_mainfunction_4"
INVOICE_DOWNLOAD_BUTTON_ID = "#btn_nouhinsho_dl_exec"
CLICKPOST_STORAGE_STATE_PATH = APP_ROOT / "data" / "storage" / "clickpost.json"
CLICKPOST_AUDIT_LOG_DIR = APP_ROOT / "logs" / "clickpost"
CLICKPOST_AUDIT_LOG_PATH = CLICKPOST_AUDIT_LOG_DIR / "clickpost_audit.jsonl"
CLICKPOST_HEADERS = (
    "お届け先郵便番号",
    "お届け先氏名",
    "お届け先敬称",
    "お届け先住所1行目",
    "お届け先住所2行目",
    "お届け先住所3行目",
    "お届け先住所4行目",
    "内容品",
)
CLICKPOST_SHIPPING_METHOD = "クリックポスト"
LETTERPACK_SHIPPING_METHOD = "レターパック500"
CLICKPOST_SHIPPING_OPTIONS = (
    "31 : レターパック500",
    "34 : クリックポスト",
)
LETTERPACK_ADDRESS_HEADERS = (
    "No",
    "宛名1（社名など）",
    "宛名2（氏名）",
    "郵便番号",
    "住所1",
    "住所2",
    "TEL",
    "品名",
    "発送方法",
    "明細行",
    "商品ｺｰﾄﾞ",
    "注文番号",
    "商品名",
    "個数",
)
CLICKPOST_TRACKING_REFLECTION_HEADERS = (
    "伝票番号",
    "発送伝票番号",
    "送り先名",
    "発送方法",
)
CLICKPOST_TRACKING_EXPORT_HEADERS = (
    "申込日時",
    "お問い合わせ番号",
    "お届け先氏名",
)
CLICKPOST_CONVERTER_BOOK_NAME = "クリックポストcsv変換.xlsm"
CLICKPOST_TRACKING_PASTE_SHEET = "送り状csv貼り付け"
CLICKPOST_TRACKING_PASTE_TABLE = "伝票番号リスト"
MEISAI_OUTPUT_TYPES = frozenset({"D_ALL", "D_KEPIN", "S_ALL", "S_KEPIN", "SETS_ALL"})
BUILDING_KEYWORDS = (
    "グラン",
    "コーポ",
    "ハイツ",
    "メゾン",
    "マンション",
    "アパート",
    "レジデンス",
    "ビル",
    "ハイム",
    "コート",
    "パレス",
    "プラザ",
    "ヴィラ",
    "シャトー",
    "シティ",
    "ライオンズ",
    "パーク",
    "カーサ",
    "ドミール",
    "ファミール",
    "サン",
    "エスポワール",
    "ラフィーネ",
    "荘",
    "寮",
)


@dataclass(frozen=True)
class ClickPostPaths:
    portal_paths: PortalPaths
    tool_root: Path
    buyer_data_dir: Path
    product_data_dir: Path
    completed_data_dir: Path
    product_list_book: Path
    content_master_book: Path


@dataclass(frozen=True)
class ContentRule:
    prefix: str
    default_quantity: int | None


@dataclass(frozen=True)
class ClickPostConversionResult:
    buyer_csv: Path
    product_csv: Path
    output_csv: Path | None
    buyer_rows: int
    product_rows: int
    target_rows: int
    output_rows: int
    warnings: tuple[str, ...]
    preview_rows: tuple[dict[str, str], ...]
    audit_path: Path | None


@dataclass(frozen=True)
class LetterPackAddressResult:
    buyer_csv: Path
    product_csv: Path
    output_csv: Path | None
    buyer_rows: int
    product_rows: int
    target_rows: int
    output_rows: int
    warnings: tuple[str, ...]
    preview_rows: tuple[dict[str, str], ...]
    audit_path: Path | None


@dataclass(frozen=True)
class ClickPostTrackingReflectionResult:
    buyer_csv: Path
    tracking_csv: Path
    output_csv: Path | None
    buyer_rows: int
    tracking_rows: int
    target_rows: int
    output_rows: int
    warnings: tuple[str, ...]
    preview_rows: tuple[dict[str, str], ...]
    audit_path: Path | None


@dataclass(frozen=True)
class ClickPostOrderListSnapshot:
    captured_at: datetime
    count: int
    order_numbers: tuple[str, ...]
    selected_shipping_options: tuple[str, ...]


@dataclass(frozen=True)
class ClickPostBuyerDownloadResult:
    executed: bool
    snapshot: ClickPostOrderListSnapshot
    downloaded_file: Path | None
    source_filename: str | None
    audit_path: Path | None
    skipped_reason: str | None


@dataclass(frozen=True)
class ClickPostProductDownloadResult:
    executed: bool
    snapshot: ClickPostOrderListSnapshot
    downloaded_file: Path | None
    source_filename: str | None
    audit_path: Path | None
    skipped_reason: str | None
    output_type: str


@dataclass(frozen=True)
class ClickPostInvoiceDownloadResult:
    executed: bool
    before_list: ClickPostOrderListSnapshot
    downloaded_file: Path | None
    source_filename: str | None
    restored: bool
    restore_result: OrderStatusBatchRestoreResult | None
    restore_verify_result: OrderStatusBatchRestoreResult | None
    skipped_reason: str | None
    error: str | None
    dialog_messages: tuple[str, ...]
    audit_path: Path | None


@dataclass(frozen=True)
class ClickPostCredential:
    yahoo_login_id: str | None
    yahoo_password: str | None
    security_code: str | None


@dataclass(frozen=True)
class ClickPostUploadResult:
    executed: bool
    csv_file: Path
    target_rows: int
    ready_for_payment: bool
    audit_path: Path | None
    skipped_reason: str | None
    warning_text: str | None


@dataclass(frozen=True)
class ClickPostPaymentPrintResult:
    executed: bool
    payment_attempts: int
    payments_completed: int
    remaining_payment_buttons: int
    print_target_rows: int
    downloaded_pdf: Path | None
    download_dir: Path
    audit_path: Path | None
    skipped_reason: str | None
    warning_text: str | None


@dataclass(frozen=True)
class ClickPostImportPaymentPrintResult:
    executed: bool
    csv_file: Path
    csv_sha256: str
    target_rows: int
    ready_for_payment: bool
    payment_attempts: int
    payments_completed: int
    remaining_payment_buttons: int
    print_target_rows: int
    downloaded_pdf: Path | None
    tracking_csv: Path | None
    tracking_rows: int
    workbook_path: Path | None
    workbook_updated: bool
    download_dir: Path
    audit_path: Path | None
    skipped_reason: str | None
    warning_text: str | None


@dataclass(frozen=True)
class ClickPostTrackingExportResult:
    executed: bool
    csv_file: Path
    target_rows: int
    tracking_rows: int
    output_csv: Path | None
    workbook_path: Path | None
    workbook_updated: bool
    audit_path: Path | None
    skipped_reason: str | None
    warning_text: str | None


@dataclass(frozen=True)
class ClickPostPreparationResult:
    buyer: ClickPostBuyerDownloadResult | None
    product: ClickPostProductDownloadResult | None
    invoice: ClickPostInvoiceDownloadResult | None
    conversion: ClickPostConversionResult
    letterpack: LetterPackAddressResult
    tracking_reflection: ClickPostTrackingReflectionResult | None
    upload: ClickPostUploadResult | None
    consistency_warnings: tuple[str, ...]
    audit_path: Path


def find_clickpost_paths() -> ClickPostPaths:
    portal_paths = find_portal_paths()
    tool_root = portal_paths.portal_root / "CP・LPP宛名作成ツール"
    paths = ClickPostPaths(
        portal_paths=portal_paths,
        tool_root=tool_root,
        buyer_data_dir=tool_root / "ネクストエンジン受注データ" / "購入者データ",
        product_data_dir=tool_root / "ネクストエンジン受注データ" / "商品情報データ",
        completed_data_dir=tool_root / "完成したデータ",
        product_list_book=tool_root / "商品リスト" / "クリックポスト対象商品リスト.xlsx",
        content_master_book=portal_paths.master_book,
    )
    if not paths.tool_root.is_dir():
        raise FileNotFoundError(f"クリックポストツールフォルダが見つかりません: {paths.tool_root}")
    return paths


def preview_clickpost_csv(
    *,
    buyer_csv: Path | None = None,
    product_csv: Path | None = None,
    preview_limit: int = 20,
) -> ClickPostConversionResult:
    return _convert_clickpost_csv(
        buyer_csv=buyer_csv,
        product_csv=product_csv,
        write=False,
        preview_limit=preview_limit,
    )


def create_clickpost_csv(
    *,
    buyer_csv: Path | None = None,
    product_csv: Path | None = None,
    preview_limit: int = 20,
) -> ClickPostConversionResult:
    return _convert_clickpost_csv(
        buyer_csv=buyer_csv,
        product_csv=product_csv,
        write=True,
        preview_limit=preview_limit,
    )


def preview_letterpack_addresses(
    *,
    buyer_csv: Path | None = None,
    product_csv: Path | None = None,
    preview_limit: int = 20,
) -> LetterPackAddressResult:
    return _build_letterpack_addresses(
        buyer_csv=buyer_csv,
        product_csv=product_csv,
        write=False,
        preview_limit=preview_limit,
    )


def create_letterpack_address_csv(
    *,
    buyer_csv: Path | None = None,
    product_csv: Path | None = None,
    preview_limit: int = 20,
) -> LetterPackAddressResult:
    return _build_letterpack_addresses(
        buyer_csv=buyer_csv,
        product_csv=product_csv,
        write=True,
        preview_limit=preview_limit,
    )


def preview_clickpost_tracking_reflection(
    *,
    tracking_csv: Path,
    buyer_csv: Path | None = None,
    preview_limit: int = 20,
) -> ClickPostTrackingReflectionResult:
    return _build_clickpost_tracking_reflection(
        tracking_csv=tracking_csv,
        buyer_csv=buyer_csv,
        write=False,
        preview_limit=preview_limit,
    )


def create_clickpost_tracking_reflection_csv(
    *,
    tracking_csv: Path,
    buyer_csv: Path | None = None,
    preview_limit: int = 20,
) -> ClickPostTrackingReflectionResult:
    return _build_clickpost_tracking_reflection(
        tracking_csv=tracking_csv,
        buyer_csv=buyer_csv,
        write=True,
        preview_limit=preview_limit,
    )

async def check_clickpost_login(
    *,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> dict[str, object]:
    client = ClickPostClient(headless=headless, slow_mo_ms=slow_mo_ms)
    return await client.check_login()


def check_clickpost_login_sync(
    *,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> dict[str, object]:
    return asyncio.run(check_clickpost_login(headless=headless, slow_mo_ms=slow_mo_ms))


async def inspect_clickpost_order_list(
    *,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostOrderListSnapshot:
    client = NextEngineClickPostClient(headless=headless, slow_mo_ms=slow_mo_ms)
    return await client.inspect_order_list()


def inspect_clickpost_order_list_sync(
    *,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostOrderListSnapshot:
    with _next_engine_storage_lock():
        return asyncio.run(inspect_clickpost_order_list(headless=headless, slow_mo_ms=slow_mo_ms))


async def download_clickpost_buyer_data(
    *,
    execute: bool,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostBuyerDownloadResult:
    client = NextEngineClickPostClient(headless=headless, slow_mo_ms=slow_mo_ms)
    return await client.download_buyer_data(execute=execute)


def download_clickpost_buyer_data_sync(
    *,
    execute: bool,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostBuyerDownloadResult:
    with _next_engine_storage_lock():
        return asyncio.run(
            download_clickpost_buyer_data(
                execute=execute,
                headless=headless,
                slow_mo_ms=slow_mo_ms,
            )
        )


async def download_clickpost_product_data(
    *,
    execute: bool,
    output_type: str = "D_ALL",
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostProductDownloadResult:
    client = NextEngineClickPostClient(headless=headless, slow_mo_ms=slow_mo_ms)
    return await client.download_product_data(execute=execute, output_type=output_type)


def download_clickpost_product_data_sync(
    *,
    execute: bool,
    output_type: str = "D_ALL",
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostProductDownloadResult:
    with _next_engine_storage_lock():
        return asyncio.run(
            download_clickpost_product_data(
                execute=execute,
                output_type=output_type,
                headless=headless,
                slow_mo_ms=slow_mo_ms,
            )
        )


async def download_clickpost_invoice_batch(
    *,
    execute: bool,
    expected_order_numbers: tuple[str, ...] = tuple(),
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostInvoiceDownloadResult:
    client = NextEngineClickPostClient(headless=headless, slow_mo_ms=slow_mo_ms)
    return await client.download_invoice_batch(
        execute=execute,
        expected_order_numbers=expected_order_numbers,
    )


def download_clickpost_invoice_batch_sync(
    *,
    execute: bool,
    expected_order_numbers: tuple[str, ...] = tuple(),
    restore_after_download: bool = False,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostInvoiceDownloadResult:
    with _next_engine_storage_lock():
        result = asyncio.run(
            download_clickpost_invoice_batch(
                execute=execute,
                expected_order_numbers=expected_order_numbers,
                headless=headless,
                slow_mo_ms=slow_mo_ms,
            )
        )

    restore_result: OrderStatusBatchRestoreResult | None = None
    restore_verify_result: OrderStatusBatchRestoreResult | None = None
    if (
        restore_after_download
        and result.executed
        and result.downloaded_file is not None
        and result.before_list.order_numbers
    ):
        restore_result = restore_next_engine_print_wait_batch_sync(
            result.before_list.order_numbers,
            execute=True,
            headless=headless,
            slow_mo_ms=slow_mo_ms,
        )
        restore_verify_result = restore_next_engine_print_wait_batch_sync(
            result.before_list.order_numbers,
            execute=False,
            headless=headless,
            slow_mo_ms=slow_mo_ms,
        )
        result = replace(
            result,
            restored=True,
            restore_result=restore_result,
            restore_verify_result=restore_verify_result,
        )

    _append_audit("invoice_download", result)
    if result.error:
        raise RuntimeError(result.error)
    return result


async def upload_clickpost_csv(
    *,
    csv_file: Path | None = None,
    execute: bool,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
    wait_at_payment_seconds: int = 0,
) -> ClickPostUploadResult:
    paths = find_clickpost_paths()
    resolved_csv = csv_file or paths.completed_data_dir / "clickpostimport.csv"
    client = ClickPostClient(headless=headless, slow_mo_ms=slow_mo_ms)
    return await client.upload_csv(
        resolved_csv,
        execute=execute,
        wait_at_payment_seconds=wait_at_payment_seconds,
    )


def upload_clickpost_csv_sync(
    *,
    csv_file: Path | None = None,
    execute: bool,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
    wait_at_payment_seconds: int = 0,
) -> ClickPostUploadResult:
    return asyncio.run(
        upload_clickpost_csv(
            csv_file=csv_file,
            execute=execute,
            headless=headless,
            slow_mo_ms=slow_mo_ms,
            wait_at_payment_seconds=wait_at_payment_seconds,
        )
    )


async def complete_clickpost_payments_and_print(
    *,
    execute: bool,
    output_dir: Path | None = None,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
    max_payments: int = 20,
) -> ClickPostPaymentPrintResult:
    paths = find_clickpost_paths()
    resolved_output_dir = output_dir or paths.completed_data_dir / "clickpost_label_pdfs"
    client = ClickPostClient(headless=headless, slow_mo_ms=slow_mo_ms)
    return await client.complete_payments_and_print(
        execute=execute,
        output_dir=resolved_output_dir,
        max_payments=max_payments,
    )


def complete_clickpost_payments_and_print_sync(
    *,
    execute: bool,
    output_dir: Path | None = None,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
    max_payments: int = 20,
) -> ClickPostPaymentPrintResult:
    return asyncio.run(
        complete_clickpost_payments_and_print(
            execute=execute,
            output_dir=output_dir,
            headless=headless,
            slow_mo_ms=slow_mo_ms,
            max_payments=max_payments,
        )
    )


async def import_pay_print_clickpost_csv(
    *,
    csv_file: Path | None = None,
    execute: bool,
    output_dir: Path | None = None,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
    max_payments: int = 20,
    progress_callback: Callable[[str, str, str | None], None] | None = None,
) -> ClickPostImportPaymentPrintResult:
    paths = find_clickpost_paths()
    resolved_csv = csv_file or paths.completed_data_dir / "clickpostimport.csv"
    resolved_output_dir = output_dir or paths.completed_data_dir / "clickpost_label_pdfs"
    client = ClickPostClient(headless=headless, slow_mo_ms=slow_mo_ms)
    return await client.import_pay_print_csv(
        resolved_csv,
        execute=execute,
        output_dir=resolved_output_dir,
        max_payments=max_payments,
        progress_callback=progress_callback,
    )


def import_pay_print_clickpost_csv_sync(
    *,
    csv_file: Path | None = None,
    execute: bool,
    output_dir: Path | None = None,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
    max_payments: int = 20,
    progress_callback: Callable[[str, str, str | None], None] | None = None,
) -> ClickPostImportPaymentPrintResult:
    return asyncio.run(
        import_pay_print_clickpost_csv(
            csv_file=csv_file,
            execute=execute,
            output_dir=output_dir,
            headless=headless,
            slow_mo_ms=slow_mo_ms,
            max_payments=max_payments,
            progress_callback=progress_callback,
        )
    )


async def export_clickpost_tracking_for_csv(
    *,
    csv_file: Path | None = None,
    execute: bool,
    output_dir: Path | None = None,
    update_workbook: bool = True,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostTrackingExportResult:
    paths = find_clickpost_paths()
    resolved_csv = csv_file or paths.completed_data_dir / "clickpostimport.csv"
    resolved_output_dir = output_dir or paths.completed_data_dir
    client = ClickPostClient(headless=headless, slow_mo_ms=slow_mo_ms)
    return await client.export_tracking_for_imported_csv(
        resolved_csv,
        execute=execute,
        output_dir=resolved_output_dir,
        update_workbook=update_workbook,
    )


def export_clickpost_tracking_for_csv_sync(
    *,
    csv_file: Path | None = None,
    execute: bool,
    output_dir: Path | None = None,
    update_workbook: bool = True,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
) -> ClickPostTrackingExportResult:
    return asyncio.run(
        export_clickpost_tracking_for_csv(
            csv_file=csv_file,
            execute=execute,
            output_dir=output_dir,
            update_workbook=update_workbook,
            headless=headless,
            slow_mo_ms=slow_mo_ms,
        )
    )


def prepare_clickpost_sync(
    *,
    fetch_next_engine: bool,
    execute_downloads: bool,
    write_conversion: bool,
    upload: bool,
    execute_upload: bool,
    output_type: str,
    headed: bool,
    slow_mo_ms: int,
    preview_limit: int,
    write_letterpack_addresses: bool = False,
    download_invoices: bool = False,
    restore_invoices_after_download: bool = False,
    tracking_csv: Path | None = None,
    write_tracking_reflection: bool = False,
    progress_callback: Callable[[str, str, str | None], None] | None = None,
) -> ClickPostPreparationResult:
    buyer: ClickPostBuyerDownloadResult | None = None
    product: ClickPostProductDownloadResult | None = None
    invoice: ClickPostInvoiceDownloadResult | None = None
    upload_result: ClickPostUploadResult | None = None

    def progress(step: str, status: str, detail: str | None = None) -> None:
        if progress_callback is None:
            return
        try:
            progress_callback(step, status, detail)
        except Exception:
            pass

    if execute_downloads:
        fetch_next_engine = True
    if execute_upload:
        upload = True
        write_conversion = True

    if fetch_next_engine:
        progress("buyer_download", "running", "購入者データを取得しています。")
        buyer = download_clickpost_buyer_data_sync(
            execute=execute_downloads,
            headless=not headed,
            slow_mo_ms=slow_mo_ms,
        )
        buyer_detail = (
            f"{buyer.snapshot.count}件"
            + (f" / {buyer.downloaded_file}" if buyer.downloaded_file else "")
        )
        progress("buyer_download", "completed", buyer_detail)

        progress("product_download", "running", "受注明細データを取得しています。")
        product = download_clickpost_product_data_sync(
            execute=execute_downloads,
            output_type=output_type,
            headless=not headed,
            slow_mo_ms=slow_mo_ms,
        )
        product_detail = (
            f"{product.snapshot.count}件"
            + (f" / {product.downloaded_file}" if product.downloaded_file else "")
        )
        progress("product_download", "completed", product_detail)

        if download_invoices:
            expected_order_numbers = buyer.snapshot.order_numbers if buyer else tuple()
            progress("invoice_download", "running", "納品書PDFをダウンロードしています。")
            invoice = download_clickpost_invoice_batch_sync(
                execute=execute_downloads,
                expected_order_numbers=expected_order_numbers,
                restore_after_download=restore_invoices_after_download,
                headless=not headed,
                slow_mo_ms=slow_mo_ms,
            )
            invoice_detail = (
                f"{invoice.before_list.count}件"
                + (f" / {invoice.downloaded_file}" if invoice.downloaded_file else "")
                + (" / 復旧済み" if invoice.restored else "")
            )
            progress("invoice_download", "completed", invoice_detail)

    downloaded_buyer_csv = buyer.downloaded_file if buyer and buyer.downloaded_file else None
    downloaded_product_csv = product.downloaded_file if product and product.downloaded_file else None

    progress(
        "clickpost_csv",
        "running",
        "クリックポストCSVを作成しています。" if write_conversion else "クリックポストCSVを確認しています。",
    )
    conversion = (
        create_clickpost_csv(
            buyer_csv=downloaded_buyer_csv,
            product_csv=downloaded_product_csv,
            preview_limit=preview_limit,
        )
        if write_conversion
        else preview_clickpost_csv(
            buyer_csv=downloaded_buyer_csv,
            product_csv=downloaded_product_csv,
            preview_limit=preview_limit,
        )
    )
    progress("clickpost_csv", "completed", f"{conversion.output_rows}件")

    progress(
        "letterpack_csv",
        "running",
        "レターパック住所CSVを作成しています。"
        if write_letterpack_addresses
        else "レターパック住所データを確認しています。",
    )
    letterpack = (
        create_letterpack_address_csv(
            buyer_csv=downloaded_buyer_csv,
            product_csv=downloaded_product_csv,
            preview_limit=preview_limit,
        )
        if write_letterpack_addresses
        else preview_letterpack_addresses(
            buyer_csv=downloaded_buyer_csv,
            product_csv=downloaded_product_csv,
            preview_limit=preview_limit,
        )
    )
    progress("letterpack_csv", "completed", f"{letterpack.output_rows}件")

    tracking_reflection = None
    if tracking_csv is not None:
        progress("tracking_reflection", "running", "クリックポスト送り状番号反映CSVを作成しています。")
        tracking_reflection = (
            create_clickpost_tracking_reflection_csv(
                tracking_csv=tracking_csv,
                preview_limit=preview_limit,
            )
            if write_tracking_reflection
            else preview_clickpost_tracking_reflection(
                tracking_csv=tracking_csv,
                preview_limit=preview_limit,
            )
        )
        progress("tracking_reflection", "completed", f"{tracking_reflection.output_rows}件")

    if upload:
        progress(
            "upload_check",
            "running",
            "クリックポストCSVをアップロード前チェックしています。" if not execute_upload else "クリックポストへCSVをアップロードしています。",
        )
        csv_file = conversion.output_csv
        if csv_file is None:
            csv_file = find_clickpost_paths().completed_data_dir / "clickpostimport.csv"
        upload_result = upload_clickpost_csv_sync(
            csv_file=csv_file,
            execute=execute_upload,
            headless=not headed,
            slow_mo_ms=slow_mo_ms,
        )
        progress("upload_check", "completed", f"{upload_result.target_rows}件")

    warnings = _preparation_warnings(buyer=buyer, product=product, invoice=invoice, conversion=conversion)
    result = ClickPostPreparationResult(
        buyer=buyer,
        product=product,
        invoice=invoice,
        conversion=conversion,
        letterpack=letterpack,
        tracking_reflection=tracking_reflection,
        upload=upload_result,
        consistency_warnings=tuple(warnings),
        audit_path=CLICKPOST_AUDIT_LOG_PATH,
    )
    _append_prepare_audit(result)
    return result


async def _new_download_context(
    browser,
    *,
    storage_state_path: Path,
    viewport_width: int = 1366,
    viewport_height: int = 900,
):
    """Create a browser context configured for ClickPost/NE download flows.

    Consolidates the ``accept_downloads`` / ``locale`` / ``viewport`` /
    ``storage_state`` context options that were previously duplicated verbatim
    in every browser session (NE order list + 4 ClickPost stages).
    """
    context_kwargs: dict[str, object] = {
        "accept_downloads": True,
        "locale": "ja-JP",
        "viewport": {"width": viewport_width, "height": viewport_height},
    }
    if storage_state_path.exists():
        context_kwargs["storage_state"] = str(storage_state_path)
    return await browser.new_context(**context_kwargs)


class NextEngineClickPostClient:
    def __init__(self, *, headless: bool | None, slow_mo_ms: int) -> None:
        self.headless = _headless_default() if headless is None else headless
        self.slow_mo_ms = slow_mo_ms
        self.paths = find_clickpost_paths()
        self.login_client = NextEngineOrderDetailDownloader(
            paths=self.paths.portal_paths,
            headless=self.headless,
            slow_mo_ms=slow_mo_ms,
        )

    async def inspect_order_list(self) -> ClickPostOrderListSnapshot:
        async with self._open_filtered_order_list() as page:
            return await _snapshot_clickpost_order_list(page)

    async def download_buyer_data(self, *, execute: bool) -> ClickPostBuyerDownloadResult:
        async with self._open_filtered_order_list() as page:
            snapshot = await _snapshot_clickpost_order_list(page)
            if not execute:
                return ClickPostBuyerDownloadResult(
                    executed=False,
                    snapshot=snapshot,
                    downloaded_file=None,
                    source_filename=None,
                    audit_path=None,
                    skipped_reason="dry_run",
                )

            if snapshot.count == 0:
                result = ClickPostBuyerDownloadResult(
                    executed=True,
                    snapshot=snapshot,
                    downloaded_file=None,
                    source_filename=None,
                    audit_path=CLICKPOST_AUDIT_LOG_PATH,
                    skipped_reason="no_orders",
                )
                _append_audit("buyer_download", result)
                return result

            destination = _next_clickpost_file_path(
                self.paths.buyer_data_dir,
                "dataクリックレター",
                ".csv",
            )
            source_filename = await _download_ne_csv_from_current_page(
                page,
                destination,
                label="clickpost_buyer",
            )

            result = ClickPostBuyerDownloadResult(
                executed=True,
                snapshot=snapshot,
                downloaded_file=destination,
                source_filename=source_filename,
                audit_path=CLICKPOST_AUDIT_LOG_PATH,
                skipped_reason=None,
            )
            _append_audit("buyer_download", result)
            return result

    async def download_product_data(
        self,
        *,
        execute: bool,
        output_type: str,
    ) -> ClickPostProductDownloadResult:
        _validate_meisai_output_type(output_type)
        async with self._open_filtered_order_list() as page:
            snapshot = await _snapshot_clickpost_order_list(page)
            if not execute:
                return ClickPostProductDownloadResult(
                    executed=False,
                    snapshot=snapshot,
                    downloaded_file=None,
                    source_filename=None,
                    audit_path=None,
                    skipped_reason="dry_run",
                    output_type=output_type,
                )

            if snapshot.count == 0:
                result = ClickPostProductDownloadResult(
                    executed=True,
                    snapshot=snapshot,
                    downloaded_file=None,
                    source_filename=None,
                    audit_path=CLICKPOST_AUDIT_LOG_PATH,
                    skipped_reason="no_orders",
                    output_type=output_type,
                )
                _append_audit("product_download", result)
                return result

            meisai_page = await _open_meisai_page(
                page,
                snapshot,
                output_type=output_type,
                login_client=self.login_client,
            )
            destination = _next_clickpost_file_path(
                self.paths.product_data_dir,
                "dataクリックレター",
                ".csv",
            )
            await _wait_for_meisai_download_link(meisai_page)
            source_filename = await _download_ne_csv_from_current_page(
                meisai_page,
                destination,
                label="clickpost_product",
            )

            result = ClickPostProductDownloadResult(
                executed=True,
                snapshot=snapshot,
                downloaded_file=destination,
                source_filename=source_filename,
                audit_path=CLICKPOST_AUDIT_LOG_PATH,
                skipped_reason=None,
                output_type=output_type,
            )
            _append_audit("product_download", result)
            return result

    async def download_invoice_batch(
        self,
        *,
        execute: bool,
        expected_order_numbers: tuple[str, ...] = tuple(),
    ) -> ClickPostInvoiceDownloadResult:
        async with self._open_filtered_order_list() as page:
            snapshot = await _snapshot_clickpost_order_list(page)
            if expected_order_numbers and snapshot.order_numbers != expected_order_numbers:
                return ClickPostInvoiceDownloadResult(
                    executed=execute,
                    before_list=snapshot,
                    downloaded_file=None,
                    source_filename=None,
                    restored=False,
                    restore_result=None,
                    restore_verify_result=None,
                    skipped_reason="target_mismatch",
                    error=(
                        "クリックポスト納品書PDFの対象伝票番号が購入者/受注明細データと一致しません。"
                        f" expected={expected_order_numbers} actual={snapshot.order_numbers}"
                    ),
                    dialog_messages=tuple(),
                    audit_path=CLICKPOST_AUDIT_LOG_PATH,
                )

            if not execute:
                return ClickPostInvoiceDownloadResult(
                    executed=False,
                    before_list=snapshot,
                    downloaded_file=None,
                    source_filename=None,
                    restored=False,
                    restore_result=None,
                    restore_verify_result=None,
                    skipped_reason="dry_run",
                    error=None,
                    dialog_messages=tuple(),
                    audit_path=None,
                )

            if snapshot.count == 0:
                return ClickPostInvoiceDownloadResult(
                    executed=True,
                    before_list=snapshot,
                    downloaded_file=None,
                    source_filename=None,
                    restored=False,
                    restore_result=None,
                    restore_verify_result=None,
                    skipped_reason="no_orders",
                    error=None,
                    dialog_messages=tuple(),
                    audit_path=CLICKPOST_AUDIT_LOG_PATH,
                )

            destination = _next_clickpost_invoice_pdf_path(self.paths)
            dialog_messages: list[str] = []
            try:
                page.on("dialog", lambda dialog: asyncio.create_task(_accept_invoice_dialog(dialog, dialog_messages)))
                source_filename = await _download_invoice_pdf_from_order_list(
                    page,
                    snapshot,
                    destination,
                )
                return ClickPostInvoiceDownloadResult(
                    executed=True,
                    before_list=snapshot,
                    downloaded_file=destination,
                    source_filename=source_filename,
                    restored=False,
                    restore_result=None,
                    restore_verify_result=None,
                    skipped_reason=None,
                    error=None,
                    dialog_messages=tuple(dialog_messages),
                    audit_path=CLICKPOST_AUDIT_LOG_PATH,
                )
            except Exception as exc:
                return ClickPostInvoiceDownloadResult(
                    executed=True,
                    before_list=snapshot,
                    downloaded_file=None,
                    source_filename=None,
                    restored=False,
                    restore_result=None,
                    restore_verify_result=None,
                    skipped_reason=None,
                    error=str(exc),
                    dialog_messages=tuple(dialog_messages),
                    audit_path=CLICKPOST_AUDIT_LOG_PATH,
                )

    def _open_filtered_order_list(self):
        return _ClickPostOrderListSession(self)


class _ClickPostOrderListSession:
    def __init__(self, client: NextEngineClickPostClient) -> None:
        self.client = client
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

    async def __aenter__(self):
        try:
            self.playwright = await async_playwright().start()
            self.browser = await self.playwright.chromium.launch(
                **_chromium_launch_options(self.client.headless, self.client.slow_mo_ms)
            )
            self.context = await _new_download_context(
                self.browser,
                storage_state_path=STORAGE_STATE_PATH,
                viewport_width=1400,
            )
            self.page = await self.context.new_page()
            _auto_accept_dialogs(self.page)
            await _open_clickpost_order_list_page(self.page, self.client.login_client)
            await self.page.wait_for_timeout(1500)
            await _filter_clickpost_shipping_methods(self.page)
            await self.context.storage_state(path=str(STORAGE_STATE_PATH))
            return self.page
        except Exception:
            await self.__aexit__(None, None, None)
            raise

    async def __aexit__(self, exc_type, exc, tb):
        if self.context is not None:
            await self.context.close()
        if self.browser is not None:
            await self.browser.close()
        if self.playwright is not None:
            await self.playwright.stop()


class ClickPostClient:
    def __init__(self, *, headless: bool | None, slow_mo_ms: int) -> None:
        self.headless = _clickpost_headless_default() if headless is None else headless
        self.slow_mo_ms = slow_mo_ms
        self.credential = _load_clickpost_credential()

    async def check_login(self) -> dict[str, object]:
        """Read-only login/session diagnostic.

        Logs in (reusing ``clickpost.json`` storage_state when still valid) and
        reports where the session landed. Performs NO CSV upload / payment /
        print side effects, so it is safe to call purely for observability and
        for verifying that the ClickPost/Yahoo login path works. Always saves a
        screenshot + HTML so a failure shows *where* it stopped.
        """
        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(
                **_chromium_launch_options(self.headless, self.slow_mo_ms)
            )
            try:
                context = await _new_download_context(
                    browser, storage_state_path=CLICKPOST_STORAGE_STATE_PATH
                )
                page = None
                try:
                    page = await context.new_page()
                    await self._login(page)
                    await page.goto(
                        CLICKPOST_MYPAGE_URL,
                        wait_until="domcontentloaded",
                        timeout=nav_timeout_ms(),
                    )
                    await page.wait_for_timeout(1000)
                    logged_in = await _page_contains(page, "マイページ", timeout=10000)
                    title = await _safe_page_title(page)
                    screenshot, html = await _save_clickpost_debug_artifacts(
                        page, "clickpost_login_check"
                    )
                    await context.storage_state(path=str(CLICKPOST_STORAGE_STATE_PATH))
                    return {
                        "logged_in": logged_in,
                        "url": getattr(page, "url", None),
                        "title": title,
                        "screenshot": str(screenshot) if screenshot else None,
                        "html": str(html) if html else None,
                    }
                finally:
                    await context.close()
            finally:
                await browser.close()

    async def upload_csv(
        self,
        csv_file: Path,
        *,
        execute: bool,
        wait_at_payment_seconds: int = 0,
    ) -> ClickPostUploadResult:
        rows = _read_csv(csv_file)
        target_rows = len(rows)
        if not execute:
            return ClickPostUploadResult(
                executed=False,
                csv_file=csv_file,
                target_rows=target_rows,
                ready_for_payment=False,
                audit_path=None,
                skipped_reason="dry_run",
                warning_text=None,
            )

        if target_rows == 0:
            result = ClickPostUploadResult(
                executed=True,
                csv_file=csv_file,
                target_rows=target_rows,
                ready_for_payment=False,
                audit_path=CLICKPOST_AUDIT_LOG_PATH,
                skipped_reason="empty_csv",
                warning_text=None,
            )
            _append_audit("upload", result)
            return result

        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(
                **_chromium_launch_options(self.headless, self.slow_mo_ms)
            )
            try:
                context = await _new_download_context(
                    browser, storage_state_path=CLICKPOST_STORAGE_STATE_PATH
                )
                try:
                    page = await context.new_page()
                    await self._login(page)
                    await self._open_multiple_create(page)
                    await page.locator('input[type="file"]').set_input_files(str(csv_file))
                    await page.wait_for_timeout(1000)
                    await _click_first_visible(
                        [
                            page.locator('input[name="commit"][type="submit"]'),
                            page.get_by_role("button", name="次へ"),
                            page.locator('input[type="submit"][value="次へ"]'),
                        ],
                        "クリックポストまとめ申込の次へ",
                    )
                    await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
                    await _click_first_visible(
                        [
                            page.locator('input[name="create"][type="submit"]'),
                            page.get_by_role("button", name="次へ"),
                            page.locator('input[type="submit"][value="次へ"]'),
                        ],
                        "クリックポストまとめ申込確認の次へ",
                    )
                    await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
                    ready = await _page_contains(page, "まとめ申込 支払手続き", timeout=60000)
                    warning_text = None if ready else await _page_text_excerpt(page)
                    await context.storage_state(path=str(CLICKPOST_STORAGE_STATE_PATH))
                    if ready and wait_at_payment_seconds > 0:
                        await page.wait_for_timeout(wait_at_payment_seconds * 1000)
                finally:
                    await context.close()
            finally:
                await browser.close()

        result = ClickPostUploadResult(
            executed=True,
            csv_file=csv_file,
            target_rows=target_rows,
            ready_for_payment=ready,
            audit_path=CLICKPOST_AUDIT_LOG_PATH,
            skipped_reason=None if ready else "payment_page_not_confirmed",
            warning_text=warning_text,
        )
        _append_audit("upload", result)
        return result

    async def complete_payments_and_print(
        self,
        *,
        execute: bool,
        output_dir: Path,
        max_payments: int,
    ) -> ClickPostPaymentPrintResult:
        output_dir.mkdir(parents=True, exist_ok=True)
        downloaded_pdf: Path | None = None
        warning_text: str | None = None
        payment_attempts = 0
        payments_completed = 0
        remaining_payment_buttons = 0
        print_target_rows = 0

        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(
                **_chromium_launch_options(self.headless, self.slow_mo_ms)
            )
            try:
                context = await _new_download_context(
                    browser, storage_state_path=CLICKPOST_STORAGE_STATE_PATH
                )
                try:
                    page = await context.new_page()
                    await self._login(page)
                    await page.goto(CLICKPOST_MYPAGE_URL, wait_until="domcontentloaded", timeout=nav_timeout_ms())
                    await page.wait_for_timeout(1500)
                    remaining_payment_buttons = await _count_visible(
                        page.locator('input.ywallet_button, input[name^="wallet_button["]')
                    )

                    if not execute:
                        await context.storage_state(path=str(CLICKPOST_STORAGE_STATE_PATH))
                        result = ClickPostPaymentPrintResult(
                            executed=False,
                            payment_attempts=0,
                            payments_completed=0,
                            remaining_payment_buttons=remaining_payment_buttons,
                            print_target_rows=0,
                            downloaded_pdf=None,
                            download_dir=output_dir,
                            audit_path=None,
                            skipped_reason="dry_run",
                            warning_text=None,
                        )
                        return result

                    while payment_attempts < max_payments:
                        payment_button = await _first_visible_locator(
                            page.locator('input.ywallet_button, input[name^="wallet_button["]'),
                            timeout=3000,
                        )
                        if payment_button is None:
                            break
                        payment_attempts += 1
                        await payment_button.click()
                        await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
                        await page.wait_for_timeout(1500)
                        await self._complete_wallet_payment(page)
                        payments_completed += 1
                        await page.goto(CLICKPOST_MYPAGE_URL, wait_until="domcontentloaded", timeout=nav_timeout_ms())
                        await page.wait_for_timeout(1500)

                    remaining_payment_buttons = await _count_visible(
                        page.locator('input.ywallet_button, input[name^="wallet_button["]')
                    )
                    if remaining_payment_buttons == 0:
                        downloaded_pdf, print_target_rows = await self._download_multiple_print_pdf(page, output_dir)
                    else:
                        warning_text = f"未決済の支払いボタンが残っています: {remaining_payment_buttons}"

                    await context.storage_state(path=str(CLICKPOST_STORAGE_STATE_PATH))
                finally:
                    await context.close()
            finally:
                await browser.close()

        result = ClickPostPaymentPrintResult(
            executed=True,
            payment_attempts=payment_attempts,
            payments_completed=payments_completed,
            remaining_payment_buttons=remaining_payment_buttons,
            print_target_rows=print_target_rows,
            downloaded_pdf=downloaded_pdf,
            download_dir=output_dir,
            audit_path=CLICKPOST_AUDIT_LOG_PATH,
            skipped_reason=None if downloaded_pdf else "pdf_not_downloaded",
            warning_text=warning_text,
        )
        _append_audit("payment_print", result)
        return result

    async def import_pay_print_csv(
        self,
        csv_file: Path,
        *,
        execute: bool,
        output_dir: Path,
        max_payments: int,
        progress_callback: Callable[[str, str, str | None], None] | None = None,
    ) -> ClickPostImportPaymentPrintResult:
        rows = _read_csv(csv_file)
        target_rows = len(rows)
        csv_sha256 = _sha256_file(csv_file)
        output_dir.mkdir(parents=True, exist_ok=True)
        ready_for_payment = False
        payment_attempts = 0
        payments_completed = 0
        remaining_payment_buttons = 0
        print_target_rows = 0
        downloaded_pdf: Path | None = None
        tracking_csv: Path | None = None
        tracking_rows = 0
        workbook_path: Path | None = None
        workbook_updated = False
        warning_text: str | None = None

        def progress(step: str, status: str, detail: str | None = None) -> None:
            if progress_callback is None:
                return
            try:
                progress_callback(step, status, detail)
            except Exception:
                pass

        progress("precheck", "running", "取込CSVを確認しています。")
        if not execute:
            progress("precheck", "completed", f"{target_rows}件 / dry_run")
            return ClickPostImportPaymentPrintResult(
                executed=False,
                csv_file=csv_file,
                csv_sha256=csv_sha256,
                target_rows=target_rows,
                ready_for_payment=False,
                payment_attempts=0,
                payments_completed=0,
                remaining_payment_buttons=0,
                print_target_rows=0,
                downloaded_pdf=None,
                tracking_csv=None,
                tracking_rows=0,
                workbook_path=None,
                workbook_updated=False,
                download_dir=output_dir,
                audit_path=None,
                skipped_reason="dry_run",
                warning_text=None,
            )

        previous_import = _find_clickpost_import_attempt(csv_sha256)
        if previous_import is not None:
            progress("precheck", "completed", "同じCSV内容の実行履歴があるため停止しました。")
            return ClickPostImportPaymentPrintResult(
                executed=False,
                csv_file=csv_file,
                csv_sha256=csv_sha256,
                target_rows=target_rows,
                ready_for_payment=False,
                payment_attempts=0,
                payments_completed=0,
                remaining_payment_buttons=0,
                print_target_rows=0,
                downloaded_pdf=None,
                tracking_csv=None,
                tracking_rows=0,
                workbook_path=None,
                workbook_updated=False,
                download_dir=output_dir,
                audit_path=CLICKPOST_AUDIT_LOG_PATH,
                skipped_reason="duplicate_csv_import_attempt",
                warning_text=(
                    "同じCSV内容のインポート実行履歴があります。"
                    f" logged_at={previous_import.get('logged_at')} kind={previous_import.get('kind')}"
                ),
            )

        progress("precheck", "completed", f"{target_rows}件")
        _append_audit(
            "import_payment_print_started",
            {
                "csv_file": csv_file,
                "csv_sha256": csv_sha256,
                "target_rows": target_rows,
            },
        )

        import_started_at = datetime.now() - timedelta(minutes=2)
        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(
                **_chromium_launch_options(self.headless, self.slow_mo_ms)
            )
            try:
                context = await _new_download_context(
                    browser, storage_state_path=CLICKPOST_STORAGE_STATE_PATH
                )
                page = None
                try:
                    page = await context.new_page()
                    progress("login", "running", "クリックポストへログインしています。")
                    await self._login(page)
                    progress("login", "completed", "ログイン完了")
                    progress("csv_import", "running", "CSVをインポートしています。")
                    await self._import_csv_to_payment_page(page, csv_file)
                    ready_for_payment = await _page_contains(page, "まとめ申込 支払手続き", timeout=60000)
                    if ready_for_payment:
                        visible_payment_buttons = await _count_visible(
                            page.locator('input.ywallet_button, input[name^="wallet_button["]')
                        )
                        if visible_payment_buttons != target_rows:
                            screenshot, html = await _save_clickpost_debug_artifacts(page, "clickpost_import_row_count_mismatch")
                            raise RuntimeError(
                                "ClickPostインポート後の支払い対象件数がCSV件数と一致しません。"
                                f" csv_rows={target_rows} payment_buttons={visible_payment_buttons}"
                                f" screenshot={screenshot} html={html}"
                            )
                        progress("csv_import", "completed", f"{visible_payment_buttons}件")
                        progress("payment", "running", "支払い手続きを実行しています。")
                        payment_attempts, payments_completed, remaining_payment_buttons = (
                            await self._complete_available_wallet_payments(page, max_payments)
                        )
                        payment_detail = f"{payments_completed}/{payment_attempts}件完了"
                        if remaining_payment_buttons:
                            payment_detail += f" / 残り{remaining_payment_buttons}件"
                        progress("payment", "completed", payment_detail)
                        if remaining_payment_buttons == 0:
                            progress("print_pdf", "running", "まとめ印字PDFを保存しています。")
                            downloaded_pdf, print_target_rows = await self._download_multiple_print_pdf(page, output_dir)
                            progress("print_pdf", "completed", f"{print_target_rows}件 / {downloaded_pdf}")
                            progress("tracking_export", "running", "送り状番号を取得してCSV/Excelへ反映しています。")
                            tracking_output_dir = (
                                output_dir.parent if output_dir.name == "clickpost_label_pdfs" else output_dir
                            )
                            tracking_csv, tracking_rows, workbook_path, workbook_updated, tracking_warnings = (
                                await self._export_imported_tracking_rows(
                                    page,
                                    rows,
                                    output_dir=tracking_output_dir,
                                    update_workbook=True,
                                    imported_after=import_started_at,
                                )
                            )
                            warning_text = _append_warning_text(warning_text, tracking_warnings)
                            progress("tracking_export", "completed", f"{tracking_rows}件 / {tracking_csv}")
                        else:
                            warning_text = f"未決済の支払いボタンが残っています: {remaining_payment_buttons}"
                    else:
                        warning_text = await _page_text_excerpt(page)
                        progress("csv_import", "completed", "支払い画面を確認できませんでした。")
                    await context.storage_state(path=str(CLICKPOST_STORAGE_STATE_PATH))
                except Exception as exc:
                    screenshot: Path | None = None
                    html: Path | None = None
                    if page is not None:
                        screenshot, html = await _save_clickpost_debug_artifacts(
                            page,
                            "clickpost_import_payment_print_error",
                        )
                    raise RuntimeError(
                        "ClickPost import/payment/print flow failed. "
                        f"url={getattr(page, 'url', None)} screenshot={screenshot} html={html}"
                    ) from exc
                finally:
                    await context.close()
            finally:
                await browser.close()

        result = ClickPostImportPaymentPrintResult(
            executed=True,
            csv_file=csv_file,
            csv_sha256=csv_sha256,
            target_rows=target_rows,
            ready_for_payment=ready_for_payment,
            payment_attempts=payment_attempts,
            payments_completed=payments_completed,
            remaining_payment_buttons=remaining_payment_buttons,
            print_target_rows=print_target_rows,
            downloaded_pdf=downloaded_pdf,
            tracking_csv=tracking_csv,
            tracking_rows=tracking_rows,
            workbook_path=workbook_path,
            workbook_updated=workbook_updated,
            download_dir=output_dir,
            audit_path=CLICKPOST_AUDIT_LOG_PATH,
            skipped_reason=None if downloaded_pdf else "pdf_not_downloaded",
            warning_text=warning_text,
        )
        _append_audit("import_payment_print", result)
        return result

    async def export_tracking_for_imported_csv(
        self,
        csv_file: Path,
        *,
        execute: bool,
        output_dir: Path,
        update_workbook: bool,
    ) -> ClickPostTrackingExportResult:
        rows = _read_csv(csv_file)
        target_rows = len(rows)
        output_dir.mkdir(parents=True, exist_ok=True)

        if not execute:
            return ClickPostTrackingExportResult(
                executed=False,
                csv_file=csv_file,
                target_rows=target_rows,
                tracking_rows=0,
                output_csv=None,
                workbook_path=None,
                workbook_updated=False,
                audit_path=None,
                skipped_reason="dry_run",
                warning_text=None,
            )

        output_csv: Path | None = None
        tracking_rows = 0
        workbook_path: Path | None = None
        workbook_updated = False
        warnings: list[str] = []

        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(
                **_chromium_launch_options(self.headless, self.slow_mo_ms)
            )
            try:
                context = await _new_download_context(
                    browser, storage_state_path=CLICKPOST_STORAGE_STATE_PATH
                )
                try:
                    page = await context.new_page()
                    await self._login(page)
                    output_csv, tracking_rows, workbook_path, workbook_updated, warnings = (
                        await self._export_imported_tracking_rows(
                            page,
                            rows,
                            output_dir=output_dir,
                            update_workbook=update_workbook,
                        )
                    )
                    await context.storage_state(path=str(CLICKPOST_STORAGE_STATE_PATH))
                finally:
                    await context.close()
            finally:
                await browser.close()

        result = ClickPostTrackingExportResult(
            executed=True,
            csv_file=csv_file,
            target_rows=target_rows,
            tracking_rows=tracking_rows,
            output_csv=output_csv,
            workbook_path=workbook_path,
            workbook_updated=workbook_updated,
            audit_path=CLICKPOST_AUDIT_LOG_PATH,
            skipped_reason=None if tracking_rows == target_rows else "tracking_rows_mismatch",
            warning_text="\n".join(warnings) if warnings else None,
        )
        _append_audit("tracking_export", result)
        return result

    async def _import_csv_to_payment_page(self, page, csv_file: Path) -> None:
        await self._open_multiple_create(page)
        await page.locator('input[type="file"]').set_input_files(str(csv_file))
        await page.wait_for_timeout(1000)
        await _click_first_visible(
            [
                page.locator('input[name="commit"][type="submit"]'),
                page.get_by_role("button", name="次へ"),
                page.locator('input[type="submit"][value="次へ"]'),
            ],
            "クリックポストまとめ申込の次へ",
        )
        await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
        await _click_first_visible(
            [
                page.locator('input[name="create"][type="submit"]'),
                page.get_by_role("button", name="次へ"),
                page.locator('input[type="submit"][value="次へ"]'),
            ],
            "クリックポストまとめ申込確認の次へ",
        )
        await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
        await page.wait_for_timeout(1500)

    async def _complete_available_wallet_payments(self, page, max_payments: int) -> tuple[int, int, int]:
        payment_attempts = 0
        payments_completed = 0
        while payment_attempts < max_payments:
            payment_button = await _first_visible_locator(
                page.locator('input.ywallet_button, input[name^="wallet_button["]'),
                timeout=3000,
            )
            if payment_button is None:
                next_payment = await _first_visible_locator(
                    page.locator(
                        "input[value*='次の支払い'], "
                        "input[value*='お支払い手続き'], "
                        "a:has-text('次の支払い'), "
                        "a:has-text('お支払い手続き'), "
                        "button:has-text('次の支払い'), "
                        "button:has-text('お支払い手続き')"
                    ),
                    timeout=1000,
                )
                if next_payment is None:
                    break
                await next_payment.click()
                await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
                await page.wait_for_timeout(1500)
                continue

            payment_attempts += 1
            await payment_button.click()
            await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
            await page.wait_for_timeout(1500)
            await self._complete_wallet_payment(page)
            payments_completed += 1
            await page.wait_for_timeout(1500)

        remaining_payment_buttons = await _count_visible(
            page.locator('input.ywallet_button, input[name^="wallet_button["]')
        )
        return payment_attempts, payments_completed, remaining_payment_buttons

    async def _export_imported_tracking_rows(
        self,
        page,
        import_rows: list[dict[str, str]],
        *,
        output_dir: Path,
        update_workbook: bool,
        imported_after: datetime | None = None,
    ) -> tuple[Path, int, Path | None, bool, list[str]]:
        mypage_rows = await self._scrape_clickpost_tracking_rows(page)
        tracking_rows, warnings = _match_imported_clickpost_tracking(
            import_rows,
            mypage_rows,
            imported_after=imported_after,
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_csv = output_dir / f"clickpost_tracking_numbers_{timestamp}.csv"
        _write_csv(output_csv, tracking_rows, CLICKPOST_TRACKING_EXPORT_HEADERS)
        complete_tracking_count = sum(1 for row in tracking_rows if _tracking_row_number(row))

        workbook_path: Path | None = None
        workbook_updated = False
        if update_workbook:
            workbook_path = _clickpost_converter_book_path()
            if complete_tracking_count != len(import_rows):
                warnings.append(
                    f"お問い合わせ番号が全件揃っていないため {CLICKPOST_CONVERTER_BOOK_NAME} は更新しませんでした: "
                    f"{complete_tracking_count}/{len(import_rows)}"
                )
            else:
                try:
                    _paste_clickpost_tracking_to_workbook(workbook_path, tracking_rows)
                    workbook_updated = True
                except PermissionError as exc:
                    warnings.append(f"{CLICKPOST_CONVERTER_BOOK_NAME} を更新できませんでした。Excelで開いている場合は閉じてください: {exc}")

        return output_csv, complete_tracking_count, workbook_path, workbook_updated, warnings

    async def _scrape_clickpost_tracking_rows(self, page) -> list[dict[str, str]]:
        await page.goto(CLICKPOST_MYPAGE_URL, wait_until="domcontentloaded", timeout=nav_timeout_ms())
        await page.wait_for_timeout(1500)
        await page.wait_for_selector("table", timeout=30000)
        rows = await page.evaluate(
            """
            () => {
              const normalize = (value) => (value || "").replace(/\\s+/g, " ").trim();
              const normalizeHeader = (value) => normalize(value).replace(/[▲▼△▽↕]/g, "");
              const tables = Array.from(document.querySelectorAll("table"));
              for (const table of tables) {
                let headerCells = Array.from(table.querySelectorAll("thead th"));
                if (headerCells.length === 0) {
                  const firstRow = table.querySelector("tr");
                  headerCells = firstRow ? Array.from(firstRow.querySelectorAll("th,td")) : [];
                }
                const headers = headerCells.map((cell) => normalizeHeader(cell.textContent));
                const indexOf = (label) => headers.findIndex((header) => header.includes(label));
                const appIndex = indexOf("申込日時");
                const inquiryIndex = indexOf("お問い合わせ番号");
                const nameIndex = indexOf("お届け先氏名");
                const contentIndex = indexOf("内容品");
                if (appIndex < 0 || inquiryIndex < 0 || nameIndex < 0) {
                  continue;
                }
                const bodyRows = Array.from(table.querySelectorAll("tbody tr"));
                const candidateRows = bodyRows.length ? bodyRows : Array.from(table.querySelectorAll("tr")).slice(1);
                return candidateRows.map((row, rowIndex) => {
                  const cells = Array.from(row.querySelectorAll("td"));
                  const cellText = (index) => index >= 0 && cells[index] ? normalize(cells[index].textContent) : "";
                  const inquiryText = cellText(inquiryIndex);
                  const inquiryMatch = inquiryText.match(/\\d{10,14}/);
                  return {
                    row_index: String(rowIndex),
                    "申込日時": cellText(appIndex),
                    "お問い合わせ番号": inquiryMatch ? inquiryMatch[0] : inquiryText,
                    "お届け先氏名": cellText(nameIndex),
                    "内容品": cellText(contentIndex),
                  };
                }).filter((item) => item["申込日時"] || item["お問い合わせ番号"] || item["お届け先氏名"]);
              }
              return [];
            }
            """
        )
        return [{str(key): str(value or "") for key, value in row.items()} for row in rows]

    async def _complete_wallet_payment(self, page) -> None:
        await self._maybe_complete_wallet_login(page)
        security_code = self.credential.security_code
        if not security_code:
            raise RuntimeError("CLICKPOST_SECURITYCODE が未設定です。")

        cvv_input = page.locator("#cvv, input[name='cvv'], input[name*='security'], input[autocomplete='cc-csc']").first
        if await _is_visible(cvv_input, timeout=15000):
            await cvv_input.fill(security_code)

        agree_checkbox = page.locator(
            "#consent-matters-agree-form-check-input, "
            "#consent-matters-agree, "
            "input[name*='agree'], "
            "input[type='checkbox']"
        )
        await _check_first_visible(agree_checkbox, "ウォレット決済規約同意")
        await _click_first_visible(
            [
                page.get_by_role("button", name="次へ"),
                page.locator("button:has-text('次へ')"),
                page.locator("input[type='submit'][value='次へ']"),
            ],
            "ウォレット決済 次へ",
        )
        confirm_button = page.locator(
            "input[type='submit'][value*='支払手続き確定'], "
            "button:has-text('支払手続き確定')"
        )
        try:
            await _first_visible_locator(confirm_button, timeout=15000)
        except Exception:
            pass
        await _click_first_visible(
            [
                page.locator("input[type='submit'][value*='支払手続き確定']"),
                page.get_by_role("button", name=re.compile("支払手続き確定")),
                page.locator("button:has-text('支払手続き確定')"),
            ],
            "支払手続き確定",
        )
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=10000)
        except PlaywrightTimeoutError:
            pass
        await page.wait_for_timeout(800)

    async def _maybe_complete_wallet_login(self, page) -> None:
        login_input = page.locator(
            "#login_handle, input[name='handle'], input[autocomplete*='username'], input[placeholder*='ID']"
        ).first
        if await _is_visible(login_input, timeout=5000):
            if self.credential.yahoo_login_id:
                await login_input.fill(self.credential.yahoo_login_id)
            await _click_first_visible(
                [
                    page.locator("button:has-text('次へ')"),
                    page.locator("button[type='submit']"),
                    page.locator("input[type='submit']"),
                ],
                "Yahooウォレットログイン次へ",
            )
            await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
            await page.wait_for_timeout(1500)

        password_input = page.locator("#password, input[name='passwd'], input[type='password']").first
        if await _is_visible(password_input, timeout=10000):
            if not self.credential.yahoo_password:
                raise RuntimeError("CLICKPOST_YAHOO_PASSWORD が未設定です。")
            password_field = password_input.first
            await password_field.fill(self.credential.yahoo_password)
            try:
                await _click_first_visible(
                    [
                        page.locator("#btnSubmit"),
                        page.locator("[name='btnSubmit']"),
                        page.locator("[name='verifyPassword']"),
                        page.get_by_role("button", name=re.compile("ログイン|同意|次へ|確認")),
                        page.locator("button:has-text('ログイン')"),
                        page.locator("button[type='submit']:not([disabled])"),
                        page.locator("input[type='submit']:not([disabled])"),
                    ],
                    "Yahooウォレットログイン実行",
                )
            except RuntimeError:
                await password_field.press("Enter")
            await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
            await page.wait_for_timeout(2500)

    async def _download_multiple_print_pdf(self, page, output_dir: Path) -> tuple[Path | None, int]:
        await page.goto(CLICKPOST_MYPAGE_URL, wait_until="domcontentloaded", timeout=nav_timeout_ms())
        await page.wait_for_timeout(1500)
        await _click_first_visible(
            [
                page.locator('input[data-url="/labels/multiple_print"]'),
                page.locator("input.navi_button[value='まとめ印字']"),
                page.locator("input[value='まとめ印字']"),
                page.get_by_role("button", name="まとめ印字"),
                page.locator("a:has-text('まとめ印字')"),
            ],
            "クリックポストまとめ印字メニュー",
        )
        await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
        await page.wait_for_timeout(1500)
        row_checkboxes = page.locator("input[type='checkbox']:not(#all_check):not(#print_agree)")
        print_target_rows = max(await row_checkboxes.count(), 0)
        if print_target_rows == 0:
            return None, 0

        all_check = page.locator("#all_check, input[name='all_check']").first
        if await _is_visible(all_check, timeout=10000):
            await all_check.check(force=True)
        else:
            for index in range(print_target_rows):
                checkbox = row_checkboxes.nth(index)
                if await checkbox.is_visible(timeout=1000):
                    await checkbox.check(force=True)

        print_agree = page.locator("#print_agree, input[name='print_agree']").first
        if await _is_visible(print_agree, timeout=10000):
            await print_agree.check(force=True)

        output_path = output_dir / f"clickpost_labels_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        download_task = asyncio.create_task(page.wait_for_event("download", timeout=30000))
        popup_task = asyncio.create_task(page.context.wait_for_event("page", timeout=30000))
        await _click_first_visible(
            [
                page.locator("#print_pdf"),
                page.locator("input[type='submit'][value='印字']"),
                page.get_by_role("button", name="印字"),
                page.locator("button:has-text('印字')"),
            ],
            "クリックポストまとめ印字",
        )
        done, pending = await asyncio.wait(
            [download_task, popup_task],
            timeout=30000,
            return_when=asyncio.FIRST_COMPLETED,
        )
        for task in pending:
            task.cancel()
        if not done:
            raise PlaywrightTimeoutError("クリックポストPDFのダウンロードまたはPDFタブを検出できませんでした。")

        value = None
        last_error: Exception | None = None
        for task in done:
            try:
                value = await task
                break
            except Exception as exc:
                last_error = exc
        if value is None:
            raise PlaywrightTimeoutError("クリックポストPDFの保存イベントがタイムアウトしました。") from last_error

        if hasattr(value, "save_as"):
            await value.save_as(str(output_path))
        else:
            await value.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
            await page.wait_for_timeout(1500)
            response = await value.context.request.get(value.url)
            content_type = response.headers.get("content-type", "")
            if not response.ok or "pdf" not in content_type.lower():
                screenshot, html = await _save_clickpost_debug_artifacts(value, "clickpost_pdf_popup_not_pdf")
                raise RuntimeError(
                    "クリックポストPDFタブを検出しましたが、PDFとして保存できませんでした。"
                    f" url={value.url} status={response.status} content_type={content_type}"
                    f" screenshot={screenshot} html={html}"
                )
            output_path.write_bytes(await response.body())
        return output_path, print_target_rows

    async def _login(self, page) -> None:
        await page.goto(CLICKPOST_URL, wait_until="domcontentloaded", timeout=nav_timeout_ms())
        await page.wait_for_timeout(1500)
        if await _page_contains(page, "マイページ", timeout=3000):
            return

        login_buttons = [
            page.locator('img[alt*="ログイン"]'),
            page.locator("a:has-text('ログイン')"),
            page.get_by_role("link", name=re.compile("ログイン")),
        ]
        await _click_first_visible(login_buttons, "クリックポストログイン")
        await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
        await page.wait_for_timeout(1500)

        login_id = self.credential.yahoo_login_id
        password = self.credential.yahoo_password
        login_input = page.locator(
            "#login_handle, input[name='handle'], input[autocomplete*='username'], input[placeholder*='ID']"
        ).first
        if await _is_visible(login_input, timeout=5000):
            if not login_id:
                raise RuntimeError("CLICKPOST_YAHOO_LOGIN_ID が未設定です。")
            await login_input.fill(login_id)
            await _click_first_visible(
                [
                    page.locator("button:has-text('次へ')"),
                    page.locator('button[type="submit"]'),
                ],
                "Yahooログイン次へ",
            )
            await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())

        password_input = page.locator("#password, input[name='passwd'], input[type='password']")
        if await _is_visible(password_input, timeout=10000):
            if not password:
                raise RuntimeError("CLICKPOST_YAHOO_PASSWORD が未設定です。")
            password_field = password_input.first
            await password_field.fill(password)
            try:
                await _click_first_visible(
                    [
                        page.locator("#btnSubmit"),
                        page.locator("[name='btnSubmit']"),
                        page.locator("[name='verifyPassword']"),
                        page.get_by_role("button", name=re.compile("ログイン|同意|次へ|確認")),
                        page.locator("button:has-text('ログイン')"),
                        page.locator('button[type="submit"]:not([disabled])'),
                        page.locator('input[type="submit"]:not([disabled])'),
                    ],
                    "Yahooログイン実行",
                )
            except RuntimeError:
                await password_field.press("Enter")
            await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
            await page.wait_for_timeout(2500)

        login_timeout = 600000 if not self.headless else 15000
        if not await _page_contains(page, "マイページ", timeout=login_timeout):
            screenshot, html = await _save_clickpost_debug_artifacts(page, "clickpost_login_failed")
            title = await _safe_page_title(page)
            excerpt = await _safe_page_excerpt(page)
            raise RuntimeError(
                "クリックポストのマイページ到達を確認できませんでした。"
                f" title={title} screenshot={screenshot} html={html} body_excerpt={excerpt}"
            )

    async def _open_multiple_create(self, page) -> None:
        await _click_first_visible(
            [
                page.locator('input[value="まとめ申込"]'),
                page.get_by_role("button", name="まとめ申込"),
                page.locator("a:has-text('まとめ申込')"),
            ],
            "クリックポストまとめ申込",
        )
        await page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
        await page.wait_for_selector('input[type="file"]', timeout=30000)


def _convert_clickpost_csv(
    *,
    buyer_csv: Path | None,
    product_csv: Path | None,
    write: bool,
    preview_limit: int,
) -> ClickPostConversionResult:
    paths = find_clickpost_paths()
    warnings: list[str] = []
    if buyer_csv is None or product_csv is None:
        found_buyer, found_product, pair_warnings = _latest_clickpost_source_pair(paths)
        buyer_csv = buyer_csv or found_buyer
        product_csv = product_csv or found_product
        warnings.extend(pair_warnings)

    buyer_rows = _read_csv(buyer_csv)
    product_rows = _read_csv(product_csv)
    products_by_order = _group_rows(product_rows, "受注番号")
    content_map = _load_content_rules(paths, warnings)

    converted: list[dict[str, str]] = []
    for buyer in buyer_rows:
        if _cell(buyer, "発送方法") != CLICKPOST_SHIPPING_METHOD:
            continue
        order_no = _cell(buyer, "受注番号")
        lines = products_by_order.get(order_no, [])
        first_line = _first_detail_line(order_no, lines, warnings)
        if first_line is None:
            continue
        address_lines = _split_clickpost_address_power_query(
            _cell(first_line, "送り先住所") or _cell(buyer, "送り先住所")
        )
        converted.append(
            {
                "お届け先郵便番号": _format_zip(_cell(first_line, "送り先〒") or _cell(buyer, "送り先〒")),
                "お届け先氏名": _cell(first_line, "送り先名") or _cell(buyer, "送り先名"),
                "お届け先敬称": "",
                "お届け先住所1行目": address_lines[0],
                "お届け先住所2行目": address_lines[1],
                "お届け先住所3行目": address_lines[2],
                "お届け先住所4行目": address_lines[3],
                "内容品": _content_for_clickpost_line(order_no, first_line, content_map, warnings),
            }
        )

    output_csv: Path | None = None
    if write:
        output_csv = paths.completed_data_dir / "clickpostimport.csv"
        _write_clickpost_csv(output_csv, converted)

    result = ClickPostConversionResult(
        buyer_csv=buyer_csv,
        product_csv=product_csv,
        output_csv=output_csv,
        buyer_rows=len(buyer_rows),
        product_rows=len(product_rows),
        target_rows=len(converted),
        output_rows=len(converted),
        warnings=tuple(dict.fromkeys(warnings)),
        preview_rows=tuple(converted[:preview_limit]),
        audit_path=CLICKPOST_AUDIT_LOG_PATH if write else None,
    )
    if write:
        _append_audit("conversion", result)
    return result


def _build_letterpack_addresses(
    *,
    buyer_csv: Path | None,
    product_csv: Path | None,
    write: bool,
    preview_limit: int,
) -> LetterPackAddressResult:
    paths = find_clickpost_paths()
    warnings: list[str] = []
    if buyer_csv is None or product_csv is None:
        found_buyer, found_product, pair_warnings = _latest_clickpost_source_pair(paths)
        buyer_csv = buyer_csv or found_buyer
        product_csv = product_csv or found_product
        warnings.extend(pair_warnings)

    buyer_rows = _read_csv(buyer_csv)
    product_rows = _read_csv(product_csv)
    products_by_order = _group_rows(product_rows, "受注番号")

    candidates = [row for row in buyer_rows if _cell(row, "発送方法") == LETTERPACK_SHIPPING_METHOD]
    candidates.sort(key=lambda row: _parse_int(_cell(row, "伝票番号")) or 0)
    candidates.sort(key=lambda row: _cell(row, "店舗"), reverse=True)

    rows: list[dict[str, str]] = []
    seen_names: set[str] = set()
    for buyer in candidates:
        order_no = _cell(buyer, "受注番号")
        product_line = _first_detail_line(order_no, products_by_order.get(order_no, []), warnings)
        if product_line is None:
            continue
        recipient = _cell(product_line, "送り先名") or _cell(buyer, "送り先名")
        if not recipient:
            warnings.append(f"レターパック対象に送り先名がありません。受注番号={order_no}")
            continue
        if recipient in seen_names:
            warnings.append(f"レターパック住所録で送り先名が重複したため2件目以降を除外しました: {recipient}")
            continue
        seen_names.add(recipient)

        address1, address2 = _split_letterpack_address_power_query(
            _cell(product_line, "送り先住所") or _cell(buyer, "送り先住所")
        )
        rows.append(
            {
                "No": str(len(rows)),
                "宛名1（社名など）": "",
                "宛名2（氏名）": recipient,
                "郵便番号": _format_zip(_cell(product_line, "送り先〒") or _cell(buyer, "送り先〒")),
                "住所1": address1,
                "住所2": address2,
                "TEL": _cell(buyer, "送り先電話番号"),
                "品名": "",
                "発送方法": LETTERPACK_SHIPPING_METHOD,
                "明細行": _cell(product_line, "明細行"),
                "商品ｺｰﾄﾞ": _cell(product_line, "商品ｺｰﾄﾞ") or _cell(product_line, "商品コード"),
                "注文番号": order_no,
                "商品名": _cell(product_line, "商品名"),
                "個数": _cell(product_line, "受注数"),
            }
        )

    output_csv: Path | None = None
    if write:
        output_csv = paths.completed_data_dir / "letterpack_addressbook.csv"
        _write_csv(output_csv, rows, LETTERPACK_ADDRESS_HEADERS)

    result = LetterPackAddressResult(
        buyer_csv=buyer_csv,
        product_csv=product_csv,
        output_csv=output_csv,
        buyer_rows=len(buyer_rows),
        product_rows=len(product_rows),
        target_rows=len(rows),
        output_rows=len(rows),
        warnings=tuple(dict.fromkeys(warnings)),
        preview_rows=tuple(rows[:preview_limit]),
        audit_path=CLICKPOST_AUDIT_LOG_PATH if write else None,
    )
    if write:
        _append_audit("letterpack_addresses", result)
    return result


def _build_clickpost_tracking_reflection(
    *,
    tracking_csv: Path,
    buyer_csv: Path | None,
    write: bool,
    preview_limit: int,
) -> ClickPostTrackingReflectionResult:
    paths = find_clickpost_paths()
    warnings: list[str] = []
    if buyer_csv is None:
        buyer_csv = _latest_data_csv(paths.buyer_data_dir)

    buyer_rows = _read_csv(buyer_csv)
    tracking_rows = _read_csv(tracking_csv)
    tracking_by_name: dict[str, list[dict[str, str]]] = {}
    for row in tracking_rows:
        name = _tracking_row_name(row)
        tracking_no = _tracking_row_number(row)
        if not name and not tracking_no:
            continue
        if not name or not tracking_no:
            warnings.append(f"クリックポスト追跡行の氏名またはお問い合わせ番号が空です: {row}")
            continue
        tracking_by_name.setdefault(name, []).append(row)

    rows: list[dict[str, str]] = []
    for buyer in buyer_rows:
        if _cell(buyer, "発送方法") != CLICKPOST_SHIPPING_METHOD:
            continue
        name = _normalize_tracking_name(_cell(buyer, "送り先名"))
        matches = tracking_by_name.get(name, [])
        if not matches:
            warnings.append(f"クリックポストお問い合わせ番号が見つかりません: 伝票番号={_cell(buyer, '伝票番号')} 送り先名={name}")
            rows.append(_tracking_reflection_row(buyer, ""))
            continue
        if len(matches) > 1:
            warnings.append(f"クリックポストお問い合わせ番号が同一氏名で複数見つかりました: {name}")
        for match in matches:
            rows.append(_tracking_reflection_row(buyer, _tracking_row_number(match)))

    output_csv: Path | None = None
    if write:
        output_csv = paths.completed_data_dir / "clickpost_tracking_reflection.csv"
        _write_csv(output_csv, rows, CLICKPOST_TRACKING_REFLECTION_HEADERS)

    result = ClickPostTrackingReflectionResult(
        buyer_csv=buyer_csv,
        tracking_csv=tracking_csv,
        output_csv=output_csv,
        buyer_rows=len(buyer_rows),
        tracking_rows=len(tracking_rows),
        target_rows=len(rows),
        output_rows=len(rows),
        warnings=tuple(dict.fromkeys(warnings)),
        preview_rows=tuple(rows[:preview_limit]),
        audit_path=CLICKPOST_AUDIT_LOG_PATH if write else None,
    )
    if write:
        _append_audit("tracking_reflection", result)
    return result


def _latest_clickpost_source_pair(paths: ClickPostPaths) -> tuple[Path, Path, list[str]]:
    buyer_csv = _latest_data_csv(paths.buyer_data_dir)
    product_csv = paths.product_data_dir / buyer_csv.name
    warnings: list[str] = []
    if not product_csv.exists():
        product_csv = _closest_timestamp_data_csv(paths.product_data_dir, buyer_csv) or _latest_data_csv(paths.product_data_dir)
        if not _looks_like_same_download_batch(buyer_csv, product_csv):
            warnings.append(
                f"購入者データと同名の商品情報データがないため、最新の商品情報データを使用しました: {product_csv}"
            )
    return buyer_csv, product_csv, warnings


def _latest_data_csv(directory: Path) -> Path:
    files = [
        path
        for path in directory.iterdir()
        if path.is_file() and path.name.lower().startswith("data") and path.suffix.lower() == ".csv"
    ]
    if not files:
        raise FileNotFoundError(f"data*.csv が見つかりません: {directory}")
    return max(files, key=lambda path: path.stat().st_mtime)


def _closest_timestamp_data_csv(directory: Path, reference: Path) -> Path | None:
    reference_time = _timestamp_from_clickpost_data_name(reference)
    if reference_time is None:
        return None
    candidates = [
        path
        for path in directory.iterdir()
        if path.is_file()
        and path.name.lower().startswith("data")
        and path.suffix.lower() == ".csv"
        and _timestamp_from_clickpost_data_name(path) is not None
    ]
    if not candidates:
        return None
    return min(
        candidates,
        key=lambda path: abs((_timestamp_from_clickpost_data_name(path) - reference_time).total_seconds()),  # type: ignore[operator]
    )


def _looks_like_same_download_batch(buyer_csv: Path, product_csv: Path) -> bool:
    buyer_time = _timestamp_from_clickpost_data_name(buyer_csv)
    product_time = _timestamp_from_clickpost_data_name(product_csv)
    if buyer_time is None or product_time is None:
        return False
    return abs((product_time - buyer_time).total_seconds()) <= 10 * 60


def _timestamp_from_clickpost_data_name(path: Path) -> datetime | None:
    match = re.search(r"(\d{10})(?=\.csv$)", path.name)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%y%m%d%H%M")
    except ValueError:
        return None


def _read_csv(path: Path) -> list[dict[str, str]]:
    for encoding in ("cp932", "utf-8-sig", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as fp:
                return [dict(row) for row in csv.DictReader(fp)]
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"CSVを読み込めません: {path}")


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fp:
        for chunk in iter(lambda: fp.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _find_clickpost_import_attempt(csv_sha256: str) -> dict[str, object] | None:
    if not CLICKPOST_AUDIT_LOG_PATH.exists():
        return None
    for line in CLICKPOST_AUDIT_LOG_PATH.read_text(encoding="utf-8").splitlines():
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue
        if payload.get("kind") not in {"import_payment_print_started", "import_payment_print"}:
            continue
        result = payload.get("result")
        if not isinstance(result, dict):
            continue
        if result.get("csv_sha256") == csv_sha256:
            return payload
    return None


def _write_clickpost_csv(path: Path, rows: Iterable[dict[str, str]]) -> None:
    _write_csv(path, rows, CLICKPOST_HEADERS)


def _write_csv(path: Path, rows: Iterable[dict[str, str]], headers: tuple[str, ...]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="cp932", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _load_content_rules(paths: ClickPostPaths, warnings: list[str]) -> dict[str, ContentRule]:
    rules: dict[str, ContentRule] = {}
    if paths.content_master_book.exists():
        _load_master_content_rules(paths.content_master_book, rules, warnings)
    if not rules:
        warnings.append("クリックポスト内容品マスタを読み込めませんでした。")
    return rules


def _first_detail_line(
    order_no: str,
    product_rows: list[dict[str, str]],
    warnings: list[str],
) -> dict[str, str] | None:
    if not product_rows:
        warnings.append(f"商品情報データに受注番号 {order_no} がありません。")
        return None
    for row in product_rows:
        if _parse_int(_cell(row, "明細行")) == 1:
            return row
    warnings.append(f"商品情報データに受注番号 {order_no} の明細行1がありません。")
    return None


def _content_for_clickpost_line(
    order_no: str,
    product_row: dict[str, str],
    rules: dict[str, ContentRule],
    warnings: list[str],
) -> str:
    code = _normalize_code(_cell(product_row, "商品ｺｰﾄﾞ") or _cell(product_row, "商品コード"))
    quantity = _cell(product_row, "受注数") or "1"
    rule = rules.get(code)
    if rule is None:
        warnings.append(f"内容品リストに商品コード {code or '(空)'} がありません。受注番号={order_no}")
        return f"{quantity})"
    return f"{rule.prefix}{quantity})"


def _split_clickpost_address_power_query(address: str) -> tuple[str, str, str, str]:
    text = unicodedata.normalize("NFKC", str(address or "")).strip()
    parts = [part.strip() for part in text.split(" ") if part.strip()]
    if len(parts) < 3:
        return _split_compact_clickpost_address(text)
    while len(parts) < 4:
        parts.append("")
    return (
        f"{parts[0]}{parts[1]}".strip(),
        parts[2].strip(),
        "".join(part.strip() for part in parts[3:]).strip(),
        "",
    )


def _split_compact_clickpost_address(address: str) -> tuple[str, str, str, str]:
    text = re.sub(r"\s+", "", address or "")
    if not text:
        return "", "", "", ""

    split_at = None
    for keyword in sorted(BUILDING_KEYWORDS, key=len, reverse=True):
        index = text.find(keyword)
        if index > 0:
            split_at = index
            break
    if split_at is None and len(text) > 20:
        split_at = 20

    if split_at is None:
        return text, "", "", ""

    first = text[:split_at]
    rest = text[split_at:]
    if len(rest) <= 20:
        return first, rest, "", ""
    return first, rest[:20], rest[20:40], rest[40:60]


def _split_letterpack_address_power_query(address: str) -> tuple[str, str]:
    text = unicodedata.normalize("NFKC", str(address or "")).strip()
    parts = [part.strip() for part in text.split(" ") if part.strip()]
    if len(parts) < 3:
        return _split_compact_letterpack_address(text)
    return (
        f"{parts[0]}{parts[1]}".strip(),
        "".join(part.strip() for part in parts[2:]).strip(),
    )


def _split_compact_letterpack_address(address: str) -> tuple[str, str]:
    text = re.sub(r"\s+", "", address or "")
    if not text:
        return "", ""

    split_at = _letterpack_room_like_start_index(text) or _building_start_index(text)
    if split_at is None and len(text) > 20:
        split_at = 20
    if split_at is None:
        return text, ""
    return text[:split_at], text[split_at:]


def _letterpack_room_like_start_index(text: str) -> int | None:
    patterns = (
        r"(?<=[0-9０-９号番\-])([ァ-ヶーA-Za-zＡ-Ｚａ-ｚ][ァ-ヶーA-Za-zＡ-Ｚａ-ｚ一-龯々〆ヵヶ0-9０-９\-]+(?:[A-Za-zＡ-Ｚａ-ｚ]?[0-9０-９]{2,}|[0-9０-９]+号室|[0-9０-９]+号).*)$",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match and match.start(1) > 5:
            return match.start(1)
    return None


def _tracking_row_name(row: dict[str, str]) -> str:
    for key in ("お届け先氏名", "Value #3", "Value#3", "value_3", "name"):
        value = _cell(row, key)
        if value:
            return _normalize_tracking_name(value)
    return ""


def _tracking_row_number(row: dict[str, str]) -> str:
    for key in ("お問い合わせ番号", "Value #2", "Value#2", "value_2", "tracking_number"):
        value = _cell(row, key)
        if value:
            return unicodedata.normalize("NFKC", value).strip()
    return ""


def _normalize_tracking_name(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "")
    text = re.sub(r"\s+", " ", text).strip()
    return text.removesuffix(" 様").removesuffix("様").strip()


def _tracking_reflection_row(buyer: dict[str, str], tracking_no: str) -> dict[str, str]:
    return {
        "伝票番号": _cell(buyer, "伝票番号"),
        "発送伝票番号": tracking_no,
        "送り先名": _cell(buyer, "送り先名"),
        "発送方法": _cell(buyer, "発送方法"),
    }


def _match_imported_clickpost_tracking(
    import_rows: list[dict[str, str]],
    mypage_rows: list[dict[str, str]],
    *,
    imported_after: datetime | None = None,
) -> tuple[list[dict[str, str]], list[str]]:
    warnings: list[str] = []
    batch_indexes = _latest_import_batch_indexes(import_rows, mypage_rows, imported_after=imported_after)
    if not batch_indexes:
        warnings.append("ClickPostマイページで今回インポート分の候補行が見つかりません。")
        if imported_after is None:
            batch_indexes = set(range(len(mypage_rows)))
    unused_indexes = {
        index
        for index, row in enumerate(mypage_rows)
        if index in batch_indexes and _tracking_row_number(row)
    }
    matched_rows: list[dict[str, str]] = []

    for import_row in import_rows:
        target_name = _normalize_tracking_name(_cell(import_row, "お届け先氏名"))
        target_content = _normalize_tracking_content(_cell(import_row, "内容品"))
        match_index = _find_tracking_match(
            mypage_rows,
            unused_indexes,
            target_name=target_name,
            target_content=target_content,
            require_content=True,
        )
        if match_index is None:
            match_index = _find_tracking_match(
                mypage_rows,
                unused_indexes,
                target_name=target_name,
                target_content=target_content,
                require_content=False,
            )
            if match_index is not None:
                warnings.append(f"内容品一致なしのため氏名のみで照合しました: {target_name}")

        if match_index is None:
            warnings.append(f"ClickPostマイページでお問い合わせ番号が見つかりません: {target_name} / {_cell(import_row, '内容品')}")
            matched_rows.append(
                {
                    "申込日時": "",
                    "お問い合わせ番号": "",
                    "お届け先氏名": _cell(import_row, "お届け先氏名"),
                }
            )
            continue

        unused_indexes.remove(match_index)
        source = mypage_rows[match_index]
        matched_rows.append(
            {
                "申込日時": _cell(source, "申込日時"),
                "お問い合わせ番号": _tracking_row_number(source),
                "お届け先氏名": _cell(source, "お届け先氏名"),
            }
        )

    return matched_rows, warnings


def _latest_import_batch_indexes(
    import_rows: list[dict[str, str]],
    mypage_rows: list[dict[str, str]],
    *,
    imported_after: datetime | None = None,
) -> set[int]:
    target_pairs = {
        (
            _normalize_tracking_name(_cell(row, "お届け先氏名")),
            _normalize_tracking_content(_cell(row, "内容品")),
        )
        for row in import_rows
    }
    candidate_indexes: list[int] = []
    parsed_dates: list[tuple[int, datetime]] = []
    for index, row in enumerate(mypage_rows):
        pair = (
            _normalize_tracking_name(_cell(row, "お届け先氏名")),
            _normalize_tracking_content(_cell(row, "内容品")),
        )
        if pair not in target_pairs:
            continue
        parsed = _parse_clickpost_application_datetime(_cell(row, "申込日時"))
        if imported_after is not None:
            if parsed is None or parsed < imported_after:
                continue
        candidate_indexes.append(index)
        if parsed is not None:
            parsed_dates.append((index, parsed))

    if not candidate_indexes:
        return set()
    if not parsed_dates:
        return set(candidate_indexes[: len(import_rows)])

    newest = max(parsed for _, parsed in parsed_dates)
    threshold = newest - timedelta(minutes=10)
    latest_indexes = {index for index, parsed in parsed_dates if parsed >= threshold}
    return latest_indexes or set(candidate_indexes[: len(import_rows)])


def _find_tracking_match(
    rows: list[dict[str, str]],
    unused_indexes: set[int],
    *,
    target_name: str,
    target_content: str,
    require_content: bool,
) -> int | None:
    for index in sorted(unused_indexes):
        row = rows[index]
        if _normalize_tracking_name(_cell(row, "お届け先氏名")) != target_name:
            continue
        if require_content and _normalize_tracking_content(_cell(row, "内容品")) != target_content:
            continue
        return index
    return None


def _normalize_tracking_content(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "")
    return re.sub(r"\s+", "", text).strip()


def _parse_clickpost_application_datetime(value: str) -> datetime | None:
    text = unicodedata.normalize("NFKC", value or "")
    text = re.sub(r"\s+", " ", text).strip()
    for fmt in ("%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _clickpost_converter_book_path() -> Path:
    return find_clickpost_paths().tool_root / CLICKPOST_CONVERTER_BOOK_NAME


def _paste_clickpost_tracking_to_workbook(workbook_path: Path, rows: list[dict[str, str]]) -> None:
    workbook = openpyxl.load_workbook(workbook_path, keep_vba=True)
    if CLICKPOST_TRACKING_PASTE_SHEET not in workbook.sheetnames:
        raise KeyError(f"{CLICKPOST_TRACKING_PASTE_SHEET} シートが見つかりません: {workbook_path}")
    worksheet = workbook[CLICKPOST_TRACKING_PASTE_SHEET]
    max_row = max(worksheet.max_row, len(rows) + 2)
    for row_index in range(1, max_row + 1):
        for column_index in range(1, 4):
            worksheet.cell(row_index, column_index).value = None

    for column_index, header in enumerate(CLICKPOST_TRACKING_EXPORT_HEADERS, start=1):
        cell = worksheet.cell(1, column_index)
        cell.value = header
        cell.number_format = "@"

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(CLICKPOST_TRACKING_EXPORT_HEADERS, start=1):
            cell = worksheet.cell(row_index, column_index)
            cell.value = _cell(row, header)
            cell.number_format = "@"

    table_last_row = max(2, len(rows) + 1)
    if CLICKPOST_TRACKING_PASTE_TABLE in worksheet.tables:
        worksheet.tables[CLICKPOST_TRACKING_PASTE_TABLE].ref = f"A1:C{table_last_row}"
    workbook.save(workbook_path)


def _append_warning_text(current: str | None, warnings: list[str]) -> str | None:
    if not warnings:
        return current
    additions = "\n".join(warnings)
    return f"{current}\n{additions}" if current else additions


def _load_product_list_rules(
    path: Path,
    rules: dict[str, ContentRule],
    warnings: list[str],
) -> None:
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        warnings.append(f"クリックポスト対象商品リストを読み込めませんでした: {exc}")
        return
    try:
        worksheet = workbook.worksheets[0]
        header = [str(value or "").strip() for value in next(worksheet.iter_rows(values_only=True))]
        code_index = _index_of(header, "商品番号")
        content_index = _index_of(header, "内容品")
        quantity_index = _index_of(header, "数量")
        if code_index is None or content_index is None:
            return
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            code = _normalize_code(row[code_index] if code_index < len(row) else "")
            prefix = str(row[content_index] or "").strip() if content_index < len(row) else ""
            quantity = _parse_int(row[quantity_index] if quantity_index is not None else None)
            if code and prefix:
                rules[code] = ContentRule(prefix=prefix, default_quantity=quantity)
    finally:
        workbook.close()


def _load_master_content_rules(
    path: Path,
    rules: dict[str, ContentRule],
    warnings: list[str],
) -> None:
    loaded, load_warnings = cached_by_mtime(
        path,
        "clickpost_content_rules",
        lambda: _read_master_content_rules_impl(path),
    )
    warnings.extend(load_warnings)
    for code, rule in loaded.items():
        if code not in rules:
            rules[code] = rule


def _read_master_content_rules_impl(
    path: Path,
) -> tuple[dict[str, ContentRule], list[str]]:
    rules: dict[str, ContentRule] = {}
    warnings: list[str] = []
    try:
        with warning_module.catch_warnings():
            warning_module.simplefilter("ignore", UserWarning)
            workbook = openpyxl.load_workbook(path, read_only=False, data_only=True, keep_vba=True)
    except Exception as exc:
        warnings.append(f"商品管理シートの内容品リストを読み込めませんでした: {exc}")
        return rules, warnings
    try:
        worksheet = workbook["クリックポスト内容品リスト"] if "クリックポスト内容品リスト" in workbook.sheetnames else workbook.worksheets[-1]
        table = worksheet.tables.get("内容品リスト")
        if table is not None:
            table_rows = [[cell.value for cell in row] for row in worksheet[table.ref]]
        else:
            table_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if not table_rows:
            return rules, warnings

        header = [str(value or "").strip() for value in table_rows[0]]
        code_index = _first_index(header, ("商品ｺｰﾄﾞ", "商品コード"))
        content_index = _index_of(header, "内容品")
        if code_index is None or content_index is None:
            return rules, warnings
        for row in table_rows[1:]:
            code = _normalize_code(row[code_index] if code_index < len(row) else "")
            prefix = str(row[content_index] or "").strip() if content_index < len(row) else ""
            if code and prefix and code not in rules:
                rules[code] = ContentRule(prefix=prefix, default_quantity=None)
    finally:
        workbook.close()
    return rules, warnings


def _content_for_order(
    order_no: str,
    product_rows: list[dict[str, str]],
    rules: dict[str, ContentRule],
    warnings: list[str],
) -> str:
    if not product_rows:
        warnings.append(f"商品情報データに伝票番号 {order_no} がありません。")
        return "雑貨"

    by_prefix: dict[str, int] = {}
    for row in product_rows:
        code = _normalize_code(_cell(row, "商品ｺｰﾄﾞ") or _cell(row, "商品コード"))
        quantity = _parse_int(_cell(row, "受注数")) or 1
        rule = rules.get(code)
        if rule is None:
            rule = ContentRule(prefix=_infer_content_prefix(_cell(row, "商品名")), default_quantity=None)
            warnings.append(f"内容品マスタに商品コード {code or '(空)'} がありません。伝票番号={order_no}")
        effective_quantity = max(quantity, rule.default_quantity or 0) if rule.default_quantity else quantity
        by_prefix[rule.prefix] = by_prefix.get(rule.prefix, 0) + max(effective_quantity, 1)

    if not by_prefix:
        return "雑貨"
    if len(by_prefix) > 1:
        warnings.append(f"伝票番号 {order_no} に複数の内容品候補があります。先頭の内容品を使用します。")
    prefix, quantity = next(iter(by_prefix.items()))
    return _format_content(prefix, quantity)


def _format_content(prefix: str, quantity: int) -> str:
    prefix = prefix.strip()
    if not prefix:
        return "雑貨"
    if "(" in prefix and not prefix.endswith(")"):
        return f"{prefix}{quantity})"
    if prefix.endswith(")"):
        return prefix
    return f"{prefix}{quantity}"


def _infer_content_prefix(product_name: str) -> str:
    if "石鹸" in product_name or "石けん" in product_name or "せっけん" in product_name:
        return "固形石鹸(PS"
    if "サプリ" in product_name or "ウコン" in product_name or "酵母" in product_name:
        return "サプリメント(SP"
    if "クリーム" in product_name:
        return "クリーム(CR"
    return "雑貨"


def _building_start_index(text: str) -> int | None:
    candidates: list[int] = []
    for keyword in BUILDING_KEYWORDS:
        index = text.find(keyword)
        if index > 5 and any(ch.isdigit() for ch in text[:index]):
            candidates.append(index)

    latin = re.search(r"(?<=[0-9])([A-Za-z][A-Za-z0-9'’\\-]*.*)$", text)
    if latin and latin.start(1) > 5:
        candidates.append(latin.start(1))

    return min(candidates) if candidates else None


def _format_zip(value: str) -> str:
    digits = re.sub(r"\D", "", unicodedata.normalize("NFKC", value or ""))
    if len(digits) == 7:
        return f"{digits[:3]}-{digits[3:]}"
    return value.strip()


def _group_rows(rows: Iterable[dict[str, str]], key: str) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        grouped.setdefault(_cell(row, key), []).append(row)
    return grouped


def _cell(row: dict[str, object], key: str) -> str:
    value = row.get(key)
    return str(value or "").strip()


def _normalize_code(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip().lower()


def _parse_int(value: object) -> int | None:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _index_of(header: list[str], name: str) -> int | None:
    return header.index(name) if name in header else None


def _first_index(header: list[str], names: tuple[str, ...]) -> int | None:
    for name in names:
        index = _index_of(header, name)
        if index is not None:
            return index
    return None


async def _open_clickpost_order_list_page(page, login_client: NextEngineOrderDetailDownloader) -> None:
    last_error = ""
    last_screenshot: Path | None = None
    last_html: Path | None = None
    for attempt in range(1, 4):
        try:
            await login_client._login(page)
            await page.goto(ORDER_LIST_PRINT_WAIT_URL, wait_until="domcontentloaded", timeout=nav_timeout_ms())
            await page.wait_for_timeout(2500)
            await login_client._remove_backdrops(page)

            if await _is_visible(page.locator("#jyuchu_dlg_open"), timeout=15000):
                return

            title = await _safe_page_title(page)
            last_error = f"search dialog button was not visible. url={page.url} title={title}"
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"

        last_screenshot, last_html = await _save_clickpost_debug_artifacts(
            page,
            f"ne_order_list_attempt_{attempt}",
        )
        if attempt < 3:
            await page.wait_for_timeout(2000)

    title = await _safe_page_title(page)
    excerpt = await _safe_page_excerpt(page)
    raise RuntimeError(
        "Next Engine order list did not become ready after login. "
        f"url={page.url} title={title} last_error={last_error} "
        f"screenshot={last_screenshot} html={last_html} body_excerpt={excerpt}"
    )


async def _safe_page_title(page) -> str:
    try:
        return await page.title()
    except Exception:
        return ""


async def _safe_page_excerpt(page) -> str:
    try:
        return await _page_text_excerpt(page)
    except Exception:
        return ""


async def _save_clickpost_debug_artifacts(page, label: str) -> tuple[Path | None, Path | None]:
    CLICKPOST_AUDIT_LOG_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = "".join(ch for ch in label if ch.isalnum() or ch in ("_", "-"))
    screenshot_path: Path | None = CLICKPOST_AUDIT_LOG_DIR / f"{timestamp}_{safe_label}.png"
    html_path: Path | None = CLICKPOST_AUDIT_LOG_DIR / f"{timestamp}_{safe_label}.html"

    try:
        await page.screenshot(path=str(screenshot_path), full_page=True)
    except Exception:
        screenshot_path = None

    try:
        html_path.write_text(await page.content(), encoding="utf-8")
    except Exception:
        html_path = None

    return screenshot_path, html_path


async def _filter_clickpost_shipping_methods(page) -> None:
    open_button = await _first_visible_locator(page.locator("#jyuchu_dlg_open"), timeout=15000)
    if open_button is None:
        screenshot, html = await _save_clickpost_debug_artifacts(page, "ne_search_dialog_button_missing")
        raise RuntimeError(
            "Next Engine search dialog button was not visible. "
            f"url={page.url} screenshot={screenshot} html={html}"
        )

    await open_button.click()
    try:
        await page.wait_for_selector("#sea_jyuchu_search_field05", timeout=30000)
    except PlaywrightTimeoutError as exc:
        screenshot, html = await _save_clickpost_debug_artifacts(page, "ne_shipping_select_missing")
        raise RuntimeError(
            "Next Engine shipping method select did not appear. "
            f"url={page.url} screenshot={screenshot} html={html}"
        ) from exc
    updated = await page.evaluate(
        """
        (wantedTexts) => {
          const normalize = (value) => (value || "").replace(/\\s+/g, " ").trim();
          const wanted = wantedTexts.map(normalize);
          const select = document.querySelector("#sea_jyuchu_search_field05");
          if (!select) return { ok: false, reason: "select_not_found" };
          const options = Array.from(select.options);
          if (!wanted.every((text) => options.some((option) => normalize(option.textContent) === text))) {
            return { ok: false, reason: "option_not_found" };
          }
          for (const option of options) {
            option.selected = wanted.includes(normalize(option.textContent));
          }
          select.dispatchEvent(new Event("change", { bubbles: true }));
          return { ok: true };
        }
        """,
        list(CLICKPOST_SHIPPING_OPTIONS),
    )
    if not updated.get("ok"):
        raise RuntimeError(f"クリックポスト/レターパックの発送方法を選択できませんでした: {updated}")

    await _click_search(page)
    await page.wait_for_timeout(3500)
    await _wait_for_clickpost_search_results(page)


async def _wait_for_clickpost_search_results(page) -> None:
    try:
        await page.wait_for_function(
            """
            () => {
              const body = document.body ? document.body.innerText : "";
              return document.querySelectorAll('input[name="qid[]"]').length > 0
                || body.includes("結果はありません")
                || body.includes("検索結果はありません");
            }
            """,
            timeout=60000,
        )
    except PlaywrightTimeoutError as exc:
        screenshot, html = await _save_clickpost_debug_artifacts(page, "ne_clickpost_search_results_missing")
        title = await _safe_page_title(page)
        excerpt = await _safe_page_excerpt(page)
        raise RuntimeError(
            "Next Engine clickpost search did not show result rows. "
            f"url={page.url} title={title} screenshot={screenshot} html={html} body_excerpt={excerpt}"
        ) from exc


async def _click_search(page) -> None:
    for selector in [
        "#ne_dlg_btn3_searchJyuchuDlg",
        "#ne_dlg_btn2_searchJyuchuDlg",
        'input[onclick="searchJyuchu.search()"]',
    ]:
        locator = page.locator(selector)
        visible_locator = await _first_visible_locator(locator, timeout=2000)
        if visible_locator is not None:
            await visible_locator.click()
            return

    screenshot, html = await _save_clickpost_debug_artifacts(page, "ne_search_button_missing")
    raise RuntimeError(
        "Next Engine search button was not visible. "
        f"url={page.url} screenshot={screenshot} html={html}"
    )
    raise RuntimeError("受注伝票管理の検索ボタンが見つかりません。")


async def _find_open_next_engine_main_page(context):
    for candidate in reversed(context.pages):
        try:
            if not candidate.is_closed() and "main.next-engine.com" in candidate.url:
                await candidate.bring_to_front()
                return candidate
        except Exception:
            continue
    return None


async def _open_next_engine_main_from_base(
    page,
    *,
    login_client: NextEngineOrderDetailDownloader | None = None,
):
    if "base.next-engine.org" not in page.url:
        return page

    main_page = await _find_open_next_engine_main_page(page.context)
    if main_page is not None:
        return main_page

    if login_client is not None:
        try:
            await login_client._login(page)
            main_page = await _find_open_next_engine_main_page(page.context)
            if main_page is not None:
                return main_page
            if "base.next-engine.org" not in page.url:
                return page
        except Exception:
            pass

    launch_locator = page.locator(
        'a[href*="/apps/launch/?id=274908"], a[href*="/apps/launch"][target="base_app_1"]'
    )
    launch = await _first_visible_locator(launch_locator, timeout=10000)
    if launch is None:
        return page

    before_pages = set(page.context.pages)
    await launch.click(force=True, timeout=15000)
    await page.wait_for_timeout(2500)

    for candidate in reversed(page.context.pages):
        if candidate in before_pages:
            continue
        try:
            if candidate.is_closed():
                continue
            await candidate.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
            await candidate.bring_to_front()
            return candidate
        except Exception:
            continue

    main_page = await _find_open_next_engine_main_page(page.context)
    return main_page or page


async def _open_meisai_page(
    page,
    snapshot: ClickPostOrderListSnapshot,
    *,
    output_type: str,
    login_client: NextEngineOrderDetailDownloader | None = None,
):
    await page.locator("#all_check").click()
    await page.wait_for_function(
        """
        (expectedCount) => {
          return document.querySelectorAll('input[name="qid[]"]:checked').length === expectedCount;
        }
        """,
        arg=snapshot.count,
        timeout=30000,
    )
    checked_order_numbers = await page.evaluate(
        """
        () => Array.from(document.querySelectorAll('input[name="qid[]"]:checked'))
          .map((element) => element.value)
          .filter(Boolean)
        """
    )
    if tuple(str(value) for value in checked_order_numbers) != snapshot.order_numbers:
        raise RuntimeError("明細一覧を開く前の選択伝票番号が検索結果と一致しません。")

    await page.locator("#extension_execute_mainfunction_6").scroll_into_view_if_needed(timeout=10000)
    await page.locator("#extension_execute_mainfunction_6 a.app_icon").click(force=True, timeout=10000)
    await page.wait_for_selector("#dialog_meisai", timeout=10000)
    await page.locator(f'#dialog_meisai input[name="type"][value="{output_type}"]').check()

    exec_button = await _first_visible_locator(page.locator("#btn_meisai_exec"), timeout=10000)
    if exec_button is None:
        screenshot, html = await _save_clickpost_debug_artifacts(page, "ne_meisai_exec_button_missing")
        raise RuntimeError(f"Next Engine detail-list execute button was not visible. screenshot={screenshot} html={html}")

    async with page.context.expect_page(timeout=60000) as new_page_info:
        await exec_button.click()
    meisai_page = await new_page_info.value
    await meisai_page.wait_for_load_state("domcontentloaded", timeout=nav_timeout_ms())
    if "base.next-engine.org" in meisai_page.url:
        meisai_page = await _open_next_engine_main_from_base(meisai_page, login_client=login_client)
        if "base.next-engine.org" in meisai_page.url:
            try:
                if meisai_page is not page:
                    await meisai_page.close()
            except Exception:
                pass
            meisai_page = await _open_next_engine_main_from_base(page, login_client=login_client)
        await meisai_page.goto(ORDER_DETAIL_LIST_URL, wait_until="domcontentloaded", timeout=nav_timeout_ms())
        await meisai_page.wait_for_timeout(2500)
    await _wait_for_meisai_download_link(meisai_page)
    return meisai_page


def _validate_meisai_output_type(output_type: str) -> None:
    if output_type not in MEISAI_OUTPUT_TYPES:
        allowed = ", ".join(sorted(MEISAI_OUTPUT_TYPES))
        raise ValueError(f"明細一覧の出力タイプが不正です: {output_type}. allowed={allowed}")


async def _wait_for_meisai_download_link(page) -> None:
    for attempt in range(1, 4):
        if "base.next-engine.org" in page.url:
            try:
                page = await _open_next_engine_main_from_base(page)
                await page.goto(ORDER_DETAIL_LIST_URL, wait_until="domcontentloaded", timeout=nav_timeout_ms())
                await page.wait_for_timeout(2500)
            except Exception:
                pass
        locator = page.locator("#searchJyuchu_table_dl_lnk")
        if await _is_visible(locator, timeout=20000):
            return
        if attempt < 3:
            try:
                await page.reload(wait_until="domcontentloaded", timeout=nav_timeout_ms())
                await page.wait_for_timeout(2500)
            except Exception:
                await page.wait_for_timeout(2500)

    screenshot, html = await _save_clickpost_debug_artifacts(page, "ne_meisai_download_link_missing")
    title = await _safe_page_title(page)
    excerpt = await _safe_page_excerpt(page)
    raise RuntimeError(
        "Next Engine detail-list download link was not visible. "
        f"url={page.url} title={title} screenshot={screenshot} html={html} body_excerpt={excerpt}"
    )


async def _download_ne_csv_from_current_page(page, destination: Path, *, label: str) -> str | None:
    locator = page.locator("#searchJyuchu_table_dl_lnk")
    last_error = ""
    for attempt in range(1, 4):
        await _prepare_next_engine_download_click(page)
        visible_locator = await _first_visible_locator(locator, timeout=20000)
        if visible_locator is None:
            last_error = "download_link_not_visible"
            if attempt < 3:
                await _reload_next_engine_download_page(page)
                continue
            break

        try:
            async with page.expect_download(timeout=download_timeout_ms(60000)) as download_info:
                await visible_locator.click(force=True)
            download = await download_info.value
            await download.save_as(str(destination))
            return download.suggested_filename
        except PlaywrightTimeoutError as exc:
            last_error = f"download_timeout:{exc}"
            if attempt < 3:
                await _reload_next_engine_download_page(page)
                continue
            break

    screenshot, html = await _save_clickpost_debug_artifacts(page, f"{label}_download_failed")
    title = await _safe_page_title(page)
    excerpt = await _safe_page_excerpt(page)
    raise RuntimeError(
        "Next Engine CSV download did not start. "
        f"label={label} url={page.url} title={title} last_error={last_error} "
        f"screenshot={screenshot} html={html} body_excerpt={excerpt}"
    )


async def _download_invoice_pdf_from_order_list(
    page,
    snapshot: ClickPostOrderListSnapshot,
    destination: Path,
) -> str | None:
    await page.locator("#all_check").click()
    await page.wait_for_function(
        """
        (expectedCount) => {
          return document.querySelectorAll('input[name="qid[]"]:checked').length === expectedCount;
        }
        """,
        arg=snapshot.count,
        timeout=30000,
    )
    checked_order_numbers = await page.evaluate(
        """
        () => Array.from(document.querySelectorAll('input[name="qid[]"]:checked'))
          .map((element) => element.value)
          .filter(Boolean)
        """
    )
    if tuple(str(value) for value in checked_order_numbers) != snapshot.order_numbers:
        raise RuntimeError("納品書PDF取得前の選択伝票番号が検索結果と一致しません。")

    await page.locator(INVOICE_ACTION_ID).scroll_into_view_if_needed(timeout=10000)
    await page.locator(INVOICE_ACTION_ID).click()
    await page.wait_for_selector(INVOICE_DOWNLOAD_BUTTON_ID, timeout=30000)
    await _set_clickpost_invoice_options(page, mode="U")

    async with page.expect_download(timeout=download_timeout_ms(180000)) as download_info:
        await page.locator(INVOICE_DOWNLOAD_BUTTON_ID).click()
    download = await download_info.value
    await download.save_as(str(destination))
    await page.context.storage_state(path=str(STORAGE_STATE_PATH))
    await page.wait_for_timeout(3000)
    return download.suggested_filename


async def _set_clickpost_invoice_options(page, *, mode: str) -> None:
    await _check_if_present(page, f'input[name="mode"][value="{mode}"]')
    await _check_if_present(page, 'input[name="ss"][value="-1"]')
    await _check_if_present(page, 'input[name="output"][value="PDF"]')


async def _check_if_present(page, selector: str) -> None:
    locator = page.locator(selector)
    if await locator.count() == 0:
        return
    try:
        await locator.first.check(timeout=5000)
    except PlaywrightTimeoutError:
        return


async def _accept_invoice_dialog(dialog, dialog_messages: list[str]) -> None:
    dialog_messages.append(dialog.message)
    await dialog.accept()


async def _prepare_next_engine_download_click(page) -> None:
    try:
        await page.keyboard.press("Escape")
    except Exception:
        pass
    try:
        await page.evaluate(
            """
            () => {
              document
                .querySelectorAll(".modal-backdrop, #cm-ov, #cc--main, .bootbox, .popover, .tooltip")
                .forEach((element) => element.remove());
              document.body.classList.remove("modal-open");
            }
            """
        )
    except Exception:
        pass
    await page.wait_for_timeout(500)


async def _reload_next_engine_download_page(page) -> None:
    try:
        await page.reload(wait_until="domcontentloaded", timeout=nav_timeout_ms())
        await page.wait_for_timeout(2500)
    except Exception:
        await page.wait_for_timeout(2500)


async def _snapshot_clickpost_order_list(page) -> ClickPostOrderListSnapshot:
    data = await page.evaluate(
        """
        () => {
          const normalize = (value) => (value || "").replace(/\\s+/g, " ").trim();
          const orderNumbers = Array.from(document.querySelectorAll('input[name="qid[]"]'))
            .map((element) => element.value)
            .filter(Boolean);
          const shippingSelect = document.querySelector("#sea_jyuchu_search_field05");
          const selectedShippingOptions = shippingSelect
            ? Array.from(shippingSelect.selectedOptions).map((option) => normalize(option.textContent))
            : [];
          return { orderNumbers, selectedShippingOptions };
        }
        """
    )
    order_numbers = tuple(str(value) for value in data["orderNumbers"])
    return ClickPostOrderListSnapshot(
        captured_at=datetime.now(),
        count=len(order_numbers),
        order_numbers=order_numbers,
        selected_shipping_options=tuple(str(value) for value in data["selectedShippingOptions"]),
    )


def _next_clickpost_file_path(directory: Path, prefix: str, suffix: str) -> Path:
    directory.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%y%m%d%H%M")
    candidate = directory / f"{prefix}{timestamp}{suffix}"
    if not candidate.exists():
        return candidate
    for index in range(1, 100):
        indexed = directory / f"{prefix}{timestamp}_{index:02d}{suffix}"
        if not indexed.exists():
            return indexed
    raise RuntimeError("保存ファイル名を決定できませんでした。")


def _next_clickpost_invoice_pdf_path(paths: ClickPostPaths) -> Path:
    directory = paths.portal_paths.portal_root / "ネクストエンジン" / "ne_納品書pdf"
    directory.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%y%m%d%H%M%S")
    candidate = directory / f"納品書_clickpost_{timestamp}.pdf"
    if not candidate.exists():
        return candidate
    for index in range(1, 100):
        indexed = directory / f"納品書_clickpost_{timestamp}_{index:02d}.pdf"
        if not indexed.exists():
            return indexed
    raise RuntimeError("納品書PDFの保存ファイル名を決定できませんでした。")


def _preparation_warnings(
    *,
    buyer: ClickPostBuyerDownloadResult | None,
    product: ClickPostProductDownloadResult | None,
    invoice: ClickPostInvoiceDownloadResult | None,
    conversion: ClickPostConversionResult,
) -> list[str]:
    warnings = list(conversion.warnings)
    if buyer and product:
        buyer_orders = set(buyer.snapshot.order_numbers)
        product_orders = set(product.snapshot.order_numbers)
        if buyer_orders != product_orders:
            warnings.append("購入者データと商品情報データの検索対象伝票番号が一致しません。")
    if buyer and invoice:
        buyer_orders = set(buyer.snapshot.order_numbers)
        invoice_orders = set(invoice.before_list.order_numbers)
        if buyer_orders and invoice_orders and buyer_orders != invoice_orders:
            warnings.append("購入者データと納品書PDFの対象伝票番号が一致しません。")
    if invoice and invoice.error:
        warnings.append(f"納品書PDFダウンロードでエラーが発生しました: {invoice.error}")
    if buyer and buyer.executed and buyer.downloaded_file:
        try:
            clickpost_rows = _count_clickpost_shipping_rows(buyer.downloaded_file)
        except Exception as exc:
            warnings.append(f"購入者データ内のクリックポスト件数を確認できませんでした: {exc}")
        else:
            if conversion.target_rows != clickpost_rows:
                warnings.append(
                    f"購入者データのクリックポスト対象 {clickpost_rows} 件に対して、CSV出力対象は {conversion.target_rows} 件です。"
                )
    return list(dict.fromkeys(warnings))


def _count_clickpost_shipping_rows(csv_path: Path) -> int:
    return sum(1 for row in _read_csv(csv_path) if _cell(row, "発送方法") == CLICKPOST_SHIPPING_METHOD)


def _load_clickpost_credential() -> ClickPostCredential:
    return ClickPostCredential(
        yahoo_login_id=os.environ.get("CLICKPOST_YAHOO_LOGIN_ID"),
        yahoo_password=os.environ.get("CLICKPOST_YAHOO_PASSWORD"),
        security_code=os.environ.get("CLICKPOST_SECURITYCODE"),
    )


def _clickpost_headless_default() -> bool:
    raw = os.environ.get("CLICKPOST_HEADLESS", "true").strip().lower()
    return raw not in {"0", "false", "no", "off"}


async def _page_contains(page, text: str, *, timeout: int) -> bool:
    try:
        await page.wait_for_function(
            "(text) => document.body && document.body.innerText.includes(text)",
            arg=text,
            timeout=timeout,
        )
        return True
    except PlaywrightTimeoutError:
        return False


async def _page_text_excerpt(page) -> str:
    text = await page.evaluate("() => (document.body && document.body.innerText || '').trim()")
    text = re.sub(r"\s+", " ", str(text))
    return text[:500]


async def _click_first_visible(locators, label: str) -> None:
    for locator in locators:
        try:
            visible_locator = await _first_visible_locator(locator, timeout=2500)
            if visible_locator is not None:
                await visible_locator.click()
                return
        except Exception:
            continue
    raise RuntimeError(f"{label} をクリックできませんでした。")


async def _check_first_visible(locator, label: str) -> None:
    visible_locator = await _first_visible_locator(locator, timeout=10000)
    if visible_locator is None:
        raise RuntimeError(f"{label} をチェックできませんでした。")
    await visible_locator.check(force=True)


async def _is_visible(locator, *, timeout: int) -> bool:
    return await _first_visible_locator(locator, timeout=timeout) is not None


async def _count_visible(locator) -> int:
    count = 0
    try:
        total = await locator.count()
        for index in range(total):
            try:
                if await locator.nth(index).is_visible(timeout=500):
                    count += 1
            except Exception:
                continue
    except Exception:
        return 0
    return count


async def _first_visible_locator(locator, *, timeout: int):
    try:
        count = await locator.count()
        if count == 0:
            return None
        if count == 1:
            return locator if await locator.is_visible(timeout=timeout) else None
        for index in range(count):
            candidate = locator.nth(index)
            if await candidate.is_visible(timeout=timeout):
                return candidate
    except PlaywrightTimeoutError:
        return None
    return None


def _append_audit(kind: str, result) -> None:
    CLICKPOST_AUDIT_LOG_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "logged_at": datetime.now().isoformat(timespec="seconds"),
        "kind": kind,
        "result": _json_safe(result),
    }
    with CLICKPOST_AUDIT_LOG_PATH.open("a", encoding="utf-8") as fp:
        fp.write(json.dumps(payload, ensure_ascii=False) + "\n")


def _append_prepare_audit(result: ClickPostPreparationResult) -> None:
    _append_audit("prepare", result)


def _json_safe(value):
    if hasattr(value, "__dataclass_fields__"):
        return _json_safe(asdict(value))
    if isinstance(value, dict):
        sanitized: dict[str, object] = {}
        for key, item in value.items():
            if key == "preview_rows":
                sanitized["preview_rows_count"] = len(item) if isinstance(item, (list, tuple)) else 0
                continue
            sanitized[str(key)] = _json_safe(item)
        return sanitized
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return _sanitize_path(value)
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    return value


def _sanitize_path(path: Path) -> str:
    raw = str(path)
    replacements = [(str(APP_ROOT), "${APP_ROOT}"), (str(Path.home()), "${USER_HOME}")]
    try:
        replacements.insert(0, (str(find_portal_paths().portal_root), "${PORTAL_ROOT}"))
    except Exception:
        pass
    for needle, replacement in replacements:
        if needle and raw.startswith(needle):
            return replacement + raw[len(needle) :]
    return raw
