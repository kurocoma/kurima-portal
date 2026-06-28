from __future__ import annotations

import csv
import json
import re
import unicodedata
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from portal_app.services.paths import PortalPaths, find_portal_paths


APP_ROOT = Path(__file__).resolve().parents[2]
YAMATO_CONVERSION_AUDIT_LOG_PATH = (
    APP_ROOT / "logs" / "next_engine_yamato" / "yamato_conversion_audit.jsonl"
)

YAMATO_INPUT_HEADERS = [
    "お客様管理番号",
    "送り状種別",
    "温度区分",
    "予備4",
    "出荷予定日",
    "配達指定日",
    "時間指定コード",
    "届け先コード",
    "届け先電話番号",
    "届け先電話番号（枝番）",
    "届け先郵便番号",
    "届け先住所",
    "届け先建物名（アパートマンション名）",
    "会社・部門名１",
    "会社・部門名２",
    "届け先名（漢字）",
    "届け先名（カナ）",
    "敬称",
    "依頼主コード",
    "依頼主電話番号",
    "依頼主電話番号（枝番）",
    "依頼主郵便番号",
    "依頼主住所",
    "依頼主建物名（アパートマンション名）",
    "依頼主名（漢字）",
    "依頼主名（カナ）",
    "品名コード１",
    "品名１",
    "品名コード２",
    "品名2",
    "荷扱い１",
    "荷扱い２",
    "記事",
    "コレクト代金引換額（税込）",
    "コレクト内消費税額",
    "営業所止置き",
    "止め置き営業所コード",
    "発行枚数",
    "個数口枠の印字",
    "請求先顧客コード",
    "請求先分類コード",
    "運賃管理番号",
]

YAMATO_OUTPUT_HEADERS = [
    "お客様管理番号",
    "送り状種別",
    "温度区分",
    "予備4",
    "出荷予定日",
    "配達指定日",
    "時間指定コード",
    "届け先コード",
    "届け先電話番号",
    "届け先電話番号（枝番）",
    "届け先郵便番号",
    "届け先住所",
    "届け先建物名（アパートマンション名）",
    "会社・部門名１",
    "会社・部門名２",
    "届け先名（漢字）",
    "届け先名（カナ）",
    "敬称",
    "依頼主コード",
    "依頼主電話番号",
    "依頼主電話番号（枝番）",
    "依頼主郵便番号",
    "依頼主住所",
    "依頼主建物名（アパートマンション名）",
    "依頼主名（漢字）",
    "依頼主名（カナ）",
    "品名コード１",
    "品名1",
    "品名コード２",
    "品名2",
    "荷扱い１",
    "荷扱い２",
    "記事",
    "コレクト代金引換額（税込）",
    "コレクト内消費税額",
    "営業所止置き",
    "止め置き営業所コード",
    "発行枚数",
    "個数口枠の印字",
    "請求先顧客コード",
    "請求先分類コード",
    "運賃管理番号",
]

PREVIEW_COLUMNS = [
    "お客様管理番号",
    "送り状種別",
    "温度区分",
    "出荷予定日",
    "配達指定日",
    "届け先郵便番号",
    "届け先住所",
    "届け先建物名（アパートマンション名）",
    "届け先名（漢字）",
    "依頼主名（漢字）",
    "品名1",
]

SOURCE_DIR_NAME = "ne-yamatocsv"
COMPLETE_DIR_NAME = "完成データ"
ITEM_NAME_TABLE = "品名テーブル_DB"
ITEM_CODE_COLUMN = "商品コード"
ITEM_NAME_COLUMN = "品名1"
DESTINATION_ADDRESS_COLUMN = "届け先住所"
DESTINATION_BUILDING_COLUMN = "届け先建物名（アパートマンション名）"
DESTINATION_NAME_COLUMN = "届け先名（漢字）"
ORDER_NO_COLUMN = "お客様管理番号"
PHONE_COLUMNS = (
    "届け先電話番号",
    "依頼主電話番号",
)
B2_ADDRESS_WIDTH_LIMIT = 64
B2_STREET_WIDTH_LIMIT = 32
B2_BUILDING_WIDTH_LIMIT = 32
HALFWIDTH_TRANSLATION = str.maketrans(
    "０１２３４５６７８９"
    "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
    "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ"
    "－−‐‑‒–—―",
    "0123456789"
    "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    "abcdefghijklmnopqrstuvwxyz"
    "--------",
)
BUILDING_KEYWORDS = (
    "マンション",
    "ハイツ",
    "コーポラス",
    "コーポ",
    "アパート",
    "レジデンス",
    "メゾン",
    "ビル",
    "タワー",
    "ヴィラ",
    "ハウス",
    "プラザ",
    "パレス",
    "テラス",
    "コート",
    "ガーデン",
    "シティ",
    "シャトー",
    "ルネ",
    "荘",
    "館",
    "寮",
    "棟",
    "号室",
    "階",
)
ENVIRONMENT_DEPENDENT_PATTERN = re.compile(r"[①-⑳Ⅰ-Ⅻ㈱㈲㈹㍉㌢㍍㌔㌘㍑㍗㌧㌦㌫㍻㋿]")
ROOM_LIKE_PATTERN = re.compile(
    r"(?:[A-Za-z]-?\d{1,4}|[A-Za-z]?-?\d{2,4}|"
    r"\d{1,4}号室|\d{1,2}階|\d{1,2}F)$"
)


@dataclass(frozen=True)
class YamatoAddressReview:
    row_number: int
    order_no: str
    destination_name: str
    original_address: str
    original_building: str
    adjusted_address: str
    adjusted_building: str
    reasons: tuple[str, ...]
    auto_adjusted: bool
    requires_review: bool


@dataclass(frozen=True)
class YamatoConversionResult:
    source_csv: Path
    master_book: Path
    output_csv: Path | None
    audit_path: Path | None
    source_rows: int
    output_rows: int
    duplicate_rows_removed: int
    item_master_rows: int
    address_adjusted_rows: int
    address_review_rows: int
    unmapped_item_codes: tuple[str, ...]
    warnings: tuple[str, ...]
    address_reviews: tuple[YamatoAddressReview, ...]
    preview_rows: tuple[dict[str, str], ...]


def preview_ne_to_yamato_conversion(
    *,
    source_csv: Path | None = None,
    preview_limit: int = 50,
) -> YamatoConversionResult:
    return convert_ne_to_yamato_csv(
        source_csv=source_csv,
        write=False,
        preview_limit=preview_limit,
    )


def create_ne_to_yamato_csv(
    *,
    source_csv: Path | None = None,
    preview_limit: int = 50,
) -> YamatoConversionResult:
    return convert_ne_to_yamato_csv(
        source_csv=source_csv,
        write=True,
        preview_limit=preview_limit,
    )


def convert_ne_to_yamato_csv(
    *,
    source_csv: Path | None = None,
    write: bool,
    preview_limit: int = 50,
) -> YamatoConversionResult:
    paths = find_portal_paths()
    ne_root = paths.portal_root / "ネクストエンジン"
    source_dir = ne_root / SOURCE_DIR_NAME
    complete_dir = ne_root / COMPLETE_DIR_NAME
    selected_source = Path(source_csv) if source_csv else latest_csv(source_dir, prefix="ne-yamato")

    item_map, item_master_rows, item_warnings = load_item_name_map(paths)
    source_df = read_yamato_source_csv(selected_source)
    (
        converted_df,
        transform_warnings,
        address_reviews,
        address_adjusted_rows,
    ) = transform_ne_to_yamato(source_df, item_map)

    output_path = None
    if write:
        complete_dir.mkdir(parents=True, exist_ok=True)
        output_path = next_output_path(complete_dir, prefix="ne-to-yamato", suffix=".csv")
        write_quoted_cp932_csv(converted_df, output_path)
    audit_path = YAMATO_CONVERSION_AUDIT_LOG_PATH if write else None

    preview_rows = dataframe_preview(converted_df, limit=preview_limit)
    unmapped = find_unmapped_item_codes(source_df, item_map)
    duplicate_removed = len(source_df) - len(converted_df)
    warnings_all = tuple(item_warnings + transform_warnings)

    result = YamatoConversionResult(
        source_csv=selected_source,
        master_book=paths.master_book,
        output_csv=output_path,
        audit_path=audit_path,
        source_rows=len(source_df),
        output_rows=len(converted_df),
        duplicate_rows_removed=duplicate_removed,
        item_master_rows=item_master_rows,
        address_adjusted_rows=address_adjusted_rows,
        address_review_rows=sum(1 for item in address_reviews if item.requires_review),
        unmapped_item_codes=tuple(unmapped),
        warnings=warnings_all,
        address_reviews=address_reviews,
        preview_rows=preview_rows,
    )
    if write:
        append_conversion_audit(result)
    return result


def transform_ne_to_yamato(
    source_df: pd.DataFrame,
    item_name_map: dict[str, str],
) -> tuple[pd.DataFrame, list[str], tuple[YamatoAddressReview, ...], int]:
    validate_headers(source_df.columns, expected=YAMATO_INPUT_HEADERS)
    warnings_out: list[str] = []

    df = source_df.copy()
    df.insert(len(df.columns), "_source_index", range(len(df)))
    df["品名１_TEMP"] = df["品名１"]
    df["品名1"] = df["品名１_TEMP"].map(item_name_map).fillna("")

    duplicate_subset = ["お客様管理番号", "届け先名（漢字）"]
    duplicate_count = int(df.duplicated(subset=duplicate_subset, keep="first").sum())
    if duplicate_count:
        warnings_out.append(
            f"{duplicate_count} 行を お客様管理番号+届け先名（漢字） の重複として除外しました。"
        )
    df = df.drop_duplicates(subset=duplicate_subset, keep="first")

    df["品名2"] = ""
    for symbol in ("★", "◆", "◎", "▲"):
        df["依頼主名（漢字）"] = df["依頼主名（漢字）"].str.replace(symbol, "", regex=False)

    phone_adjusted_rows = apply_b2_phone_rules(df)
    if phone_adjusted_rows:
        warnings_out.append(
            f"{phone_adjusted_rows} 行の電話番号をB2向けに自動補正しました。"
        )

    address_reviews, address_adjusted_rows = apply_b2_address_rules(df)
    address_review_rows = sum(1 for item in address_reviews if item.requires_review)
    if address_adjusted_rows:
        warnings_out.append(
            f"{address_adjusted_rows} 行の届け先住所/建物名をB2仕様に合わせて自動補正しました。"
        )
    if address_review_rows:
        warnings_out.append(
            f"{address_review_rows} 行の届け先住所/建物名は手動確認が必要です。"
        )

    missing_item_codes = find_unmapped_item_codes(source_df, item_name_map)
    if missing_item_codes:
        warnings_out.append(
            "品名変換マスタに存在しない品名コードがあります: "
            + ", ".join(missing_item_codes[:20])
            + (" ..." if len(missing_item_codes) > 20 else "")
        )

    df = df.sort_values("_source_index", kind="stable")
    df = df[YAMATO_OUTPUT_HEADERS].fillna("").astype(str)
    return df, warnings_out, tuple(address_reviews), address_adjusted_rows


def apply_b2_phone_rules(dataframe: pd.DataFrame) -> int:
    adjusted_indexes: set[object] = set()
    for column in PHONE_COLUMNS:
        if column not in dataframe.columns:
            continue
        for index, value in dataframe[column].items():
            original = normalize_cell(value)
            adjusted = normalize_b2_phone_number(original)
            if adjusted != original:
                dataframe.at[index, column] = adjusted
                adjusted_indexes.add(index)
    return len(adjusted_indexes)


def normalize_b2_phone_number(value: object) -> str:
    text = normalize_cell(value).translate(HALFWIDTH_TRANSLATION)
    text = text.replace("＋", "+")
    if not text:
        return ""

    digits = re.sub(r"\D", "", text)
    if text.startswith("+81"):
        national = digits[2:]
        if national and not national.startswith("0"):
            return "0" + national
        return national

    if text.startswith("0081"):
        national = digits[4:]
        if national and not national.startswith("0"):
            return "0" + national
        return national

    return digits


def apply_b2_address_rules(dataframe: pd.DataFrame) -> tuple[list[YamatoAddressReview], int]:
    reviews: list[YamatoAddressReview] = []
    adjusted_rows = 0

    for index, row in dataframe.iterrows():
        original_address = normalize_cell(row.get(DESTINATION_ADDRESS_COLUMN))
        original_building = normalize_cell(row.get(DESTINATION_BUILDING_COLUMN))
        adjusted_address, adjusted_building, reasons, requires_review = split_b2_address(
            original_address,
            original_building,
        )

        auto_adjusted = (
            adjusted_address != original_address
            or adjusted_building != original_building
        )
        if auto_adjusted:
            dataframe.at[index, DESTINATION_ADDRESS_COLUMN] = adjusted_address
            dataframe.at[index, DESTINATION_BUILDING_COLUMN] = adjusted_building
            adjusted_rows += 1

        if auto_adjusted or requires_review:
            reviews.append(
                YamatoAddressReview(
                    row_number=int(row.get("_source_index", index)) + 2,
                    order_no=normalize_cell(row.get(ORDER_NO_COLUMN)),
                    destination_name=normalize_cell(row.get(DESTINATION_NAME_COLUMN)),
                    original_address=original_address,
                    original_building=original_building,
                    adjusted_address=adjusted_address,
                    adjusted_building=adjusted_building,
                    reasons=tuple(reasons),
                    auto_adjusted=auto_adjusted,
                    requires_review=requires_review,
                )
            )

    return reviews, adjusted_rows


def split_b2_address(
    original_address: str,
    original_building: str,
) -> tuple[str, str, list[str], bool]:
    reasons: list[str] = []
    requires_review = False

    address = normalize_b2_text(original_address)
    building = normalize_b2_text(original_building)
    if address != original_address or building != original_building:
        reasons.append("数字・英字・ハイフン・空白をB2向けに正規化")

    if building:
        structure_changed = False
        split_index, split_reason = find_address_building_split(address, building)
        if split_index is None:
            adjusted_address = remove_address_spaces(address)
            adjusted_building = remove_address_spaces(building)
        else:
            structure_changed = True
            moved_building_part = address[split_index:]
            adjusted_address = remove_address_spaces(address[:split_index])
            adjusted_building = remove_address_spaces(moved_building_part + building)
            reasons.append(split_reason)

        if should_join_short_building_to_address(adjusted_address, adjusted_building):
            structure_changed = True
            adjusted_address = adjusted_address + adjusted_building
            adjusted_building = ""
            reasons.append("1桁の建物名欄を住所末尾に結合")

        (
            adjusted_address,
            adjusted_building,
            overflow_reason,
        ) = move_building_prefix_to_street_if_needed(
            adjusted_address,
            adjusted_building,
        )
        if overflow_reason:
            structure_changed = True
            reasons.append(overflow_reason)

        if not structure_changed and adjusted_address != address:
            reasons.append("住所欄の空白を除去")
        if not structure_changed and adjusted_building != building:
            reasons.append("建物名欄の空白を除去")
    else:
        split_index, split_reason = find_address_building_split(address)
        if split_index is None:
            adjusted_address = remove_address_spaces(address)
            adjusted_building = ""
            if adjusted_address != address:
                reasons.append("住所欄の空白を除去")
        else:
            adjusted_address = remove_address_spaces(address[:split_index])
            adjusted_building = remove_address_spaces(address[split_index:])
            reasons.append(split_reason)

    if not adjusted_address:
        requires_review = True
        reasons.append("届け先住所が空欄")
    if b2_text_width(adjusted_address) > B2_ADDRESS_WIDTH_LIMIT:
        requires_review = True
        reasons.append(
            f"届け先住所がB2上限{B2_ADDRESS_WIDTH_LIMIT}幅を超過"
        )
    if b2_text_width(b2_street_component(adjusted_address)) > B2_STREET_WIDTH_LIMIT:
        requires_review = True
        reasons.append(
            f"届け先町・番地がB2上限{B2_STREET_WIDTH_LIMIT}幅を超過"
        )
    if b2_text_width(adjusted_building) > B2_BUILDING_WIDTH_LIMIT:
        requires_review = True
        reasons.append(
            f"届け先建物名がB2上限{B2_BUILDING_WIDTH_LIMIT}幅を超過"
        )
    if contains_environment_dependent_chars(adjusted_address + adjusted_building):
        requires_review = True
        reasons.append("環境依存文字の可能性")
    if not can_encode_cp932(adjusted_address + adjusted_building):
        requires_review = True
        reasons.append("CP932で出力できない文字を含む")
    if re.search(r"[.．]", adjusted_address):
        requires_review = True
        reasons.append("住所欄にドットが残存")

    return adjusted_address, adjusted_building, dedupe_preserve_order(reasons), requires_review


def find_address_building_split(
    address: str,
    existing_building: str = "",
) -> tuple[int | None, str | None]:
    if not address:
        return None, None

    for match in re.finditer(r"\s+", address):
        before = address[: match.start()]
        after = address[match.end() :]
        if before and after and looks_like_building_candidate(after, existing_building):
            return match.start(), "住所内の空白以降を建物名欄へ分割"

    candidates: list[int] = []
    street_number_pattern = re.compile(
        r"(?:\d+丁目|\d+番地?|\d+号|\d+(?:-\d+)+|\d+)"
    )
    for match in street_number_pattern.finditer(address):
        suffix = address[match.end() :]
        if suffix and looks_like_building_candidate(suffix, existing_building):
            candidates.append(match.end())

    if candidates:
        return max(candidates), "町番地以降の建物名らしい文字列を建物名欄へ移動"

    return None, None


def looks_like_building_candidate(value: str, existing_building: str = "") -> bool:
    return looks_like_building(value) or (
        bool(existing_building) and looks_like_building(value + existing_building)
    )


def should_join_short_building_to_address(address: str, building: str) -> bool:
    if not re.fullmatch(r"\d", building):
        return False
    if not re.search(r"(?:-\d+|\d+番地?|\d+号|\d+丁目|\d+)$", address):
        return False
    if b2_text_width(address + building) > B2_ADDRESS_WIDTH_LIMIT:
        return False
    return b2_text_width(b2_street_component(address) + building) <= B2_STREET_WIDTH_LIMIT


def move_building_prefix_to_street_if_needed(
    address: str,
    building: str,
) -> tuple[str, str, str | None]:
    if not building or b2_text_width(building) <= B2_BUILDING_WIDTH_LIMIT:
        return address, building, None

    street = b2_street_component(address)
    available_width = B2_STREET_WIDTH_LIMIT - b2_text_width(street)
    if available_width <= 0:
        return address, building, None

    split_index = find_building_prefix_split(building, max_prefix_width=available_width)
    if split_index is None:
        return address, building, None

    moved_prefix = building[:split_index]
    remaining_building = building[split_index:]
    if not moved_prefix or not remaining_building:
        return address, building, None

    adjusted_address = address + moved_prefix
    adjusted_building = remaining_building
    return (
        adjusted_address,
        adjusted_building,
        "建物名が長いため町・番地の空き文字数に収まる範囲を住所欄へ移動",
    )


def find_building_prefix_split(building: str, *, max_prefix_width: int) -> int | None:
    valid_indices = [
        index
        for index in range(1, len(building))
        if b2_text_width(building[:index]) <= max_prefix_width
        and b2_text_width(building[index:]) <= B2_BUILDING_WIDTH_LIMIT
    ]
    if not valid_indices:
        return None

    semantic_indices: set[int] = set()
    for keyword in BUILDING_KEYWORDS:
        for match in re.finditer(re.escape(keyword), building):
            semantic_indices.add(match.end())

    for match in re.finditer(r"(?=\d{1,4}(?:号室?|室|階|F|[A-Za-z]|-))", building):
        semantic_indices.add(match.start())

    semantic_candidates = [index for index in valid_indices if index in semantic_indices]
    if semantic_candidates:
        return max(semantic_candidates, key=lambda index: b2_text_width(building[:index]))
    return max(valid_indices, key=lambda index: b2_text_width(building[:index]))


def b2_street_component(address: str) -> str:
    text = remove_address_spaces(address)
    if not text:
        return ""

    first_digit = re.search(r"\d", text)
    search_area = text[: first_digit.start()] if first_digit else text
    marker_index = max(search_area.rfind(marker) for marker in ("市", "区", "町", "村"))
    if marker_index >= 0:
        return text[marker_index + 1 :]
    return text


def looks_like_building(value: str) -> bool:
    text = remove_address_spaces(value)
    if not text:
        return False
    return (
        any(keyword in text for keyword in BUILDING_KEYWORDS)
        or looks_like_care_of(text)
    ) or bool(
        ROOM_LIKE_PATTERN.search(text)
    )


def looks_like_care_of(value: str) -> bool:
    text = remove_address_spaces(value)
    if b2_text_width(text) > B2_BUILDING_WIDTH_LIMIT:
        return False
    if text.endswith("様方") and len(text) > 2:
        return True
    if text.endswith("方") and len(text) > 1:
        return not text.endswith(("地方", "平方", "方向"))
    return False


def normalize_b2_text(value: str) -> str:
    text = normalize_cell(value).translate(HALFWIDTH_TRANSLATION)
    text = text.replace("　", " ")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s*-\s*", "-", text)
    return text.strip()


def remove_address_spaces(value: str) -> str:
    return re.sub(r"[ 　]+", "", value)


def b2_text_width(value: str) -> int:
    width = 0
    for char in value:
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W", "A"} else 1
    return width


def contains_environment_dependent_chars(value: str) -> bool:
    return bool(ENVIRONMENT_DEPENDENT_PATTERN.search(value))


def can_encode_cp932(value: str) -> bool:
    try:
        value.encode("cp932")
        return True
    except UnicodeEncodeError:
        return False


def dedupe_preserve_order(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def read_yamato_source_csv(source_csv: Path) -> pd.DataFrame:
    if not source_csv.is_file():
        raise FileNotFoundError(f"ne-yamato CSV が見つかりません: {source_csv}")
    return pd.read_csv(
        source_csv,
        encoding="cp932",
        dtype=str,
        keep_default_na=False,
        na_filter=False,
    )


def load_item_name_map(paths: PortalPaths) -> tuple[dict[str, str], int, list[str]]:
    rows = read_excel_table(paths.master_book, ITEM_NAME_TABLE)
    warnings_out: list[str] = []
    mapping: dict[str, str] = {}
    duplicate_codes: set[str] = set()

    for row in rows:
        code = normalize_cell(row.get(ITEM_CODE_COLUMN))
        item_name = normalize_cell(row.get(ITEM_NAME_COLUMN))
        if not code:
            continue
        if code in mapping:
            duplicate_codes.add(code)
            continue
        mapping[code] = item_name

    if duplicate_codes:
        warnings_out.append(
            f"{ITEM_NAME_TABLE} に重複した商品コードがあります: "
            + ", ".join(sorted(duplicate_codes)[:20])
        )

    return mapping, len(rows), warnings_out


def read_excel_table(workbook_path: Path, table_name: str) -> list[dict[str, object]]:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"商品管理シートが見つかりません: {workbook_path}")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(workbook_path, read_only=False, data_only=True, keep_vba=False)

    try:
        for worksheet in workbook.worksheets:
            table = worksheet.tables.get(table_name)
            if table is None:
                continue

            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            headers = [
                worksheet.cell(min_row, column).value
                for column in range(min_col, max_col + 1)
            ]
            rows: list[dict[str, object]] = []
            for row_number in range(min_row + 1, max_row + 1):
                values = [
                    worksheet.cell(row_number, column).value
                    for column in range(min_col, max_col + 1)
                ]
                if any(value not in (None, "") for value in values):
                    rows.append(dict(zip(headers, values, strict=False)))
            return rows
    finally:
        workbook.close()

    raise ValueError(f"Excelテーブル {table_name} が見つかりません: {workbook_path}")


def validate_headers(actual: Iterable[str], *, expected: list[str]) -> None:
    actual_list = list(actual)
    if actual_list == expected:
        return

    missing = [column for column in expected if column not in actual_list]
    extra = [column for column in actual_list if column not in expected]
    detail = []
    if missing:
        detail.append("不足=" + ", ".join(missing))
    if extra:
        detail.append("余分=" + ", ".join(extra))
    raise ValueError("ne-yamato CSV のヘッダーが想定と一致しません。" + " / ".join(detail))


def find_unmapped_item_codes(source_df: pd.DataFrame, item_name_map: dict[str, str]) -> list[str]:
    if "品名１" not in source_df.columns:
        return []

    codes = []
    seen: set[str] = set()
    for value in source_df["品名１"].fillna("").astype(str):
        code = value.strip()
        if code and code not in item_name_map and code not in seen:
            seen.add(code)
            codes.append(code)
    return codes


def latest_csv(directory: Path, *, prefix: str) -> Path:
    if not directory.is_dir():
        raise FileNotFoundError(f"CSVフォルダが見つかりません: {directory}")
    files = [
        path
        for path in directory.iterdir()
        if path.is_file()
        and path.suffix.lower() == ".csv"
        and path.name.lower().startswith(prefix.lower())
    ]
    if not files:
        raise FileNotFoundError(f"{prefix} で始まるCSVが見つかりません: {directory}")
    return max(files, key=lambda path: path.stat().st_mtime)


def next_output_path(directory: Path, *, prefix: str, suffix: str) -> Path:
    timestamp = datetime.now().strftime("%y%m%d%H%M")
    candidate = directory / f"{prefix}{timestamp}{suffix}"
    if not candidate.exists():
        return candidate

    for index in range(1, 100):
        indexed = directory / f"{prefix}{timestamp}_{index:02d}{suffix}"
        if not indexed.exists():
            return indexed
    raise RuntimeError(f"出力ファイル名を決定できません: {directory}")


def write_quoted_cp932_csv(dataframe: pd.DataFrame, output_path: Path) -> None:
    with output_path.open("w", encoding="cp932", newline="") as fp:
        writer = csv.writer(fp, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        writer.writerow(YAMATO_OUTPUT_HEADERS)
        for row in dataframe.itertuples(index=False, name=None):
            writer.writerow(["" if value is None else str(value) for value in row])


def append_conversion_audit(result: YamatoConversionResult) -> None:
    if result.audit_path is None:
        return

    result.audit_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "logged_at": datetime.now().isoformat(timespec="seconds"),
        "kind": "ne_to_yamato_conversion",
        "source_csv": str(result.source_csv),
        "master_book": str(result.master_book),
        "output_csv": str(result.output_csv) if result.output_csv else None,
        "source_rows": result.source_rows,
        "output_rows": result.output_rows,
        "duplicate_rows_removed": result.duplicate_rows_removed,
        "item_master_rows": result.item_master_rows,
        "address_adjusted_rows": result.address_adjusted_rows,
        "address_review_rows": result.address_review_rows,
        "address_reviews": [
            {
                "row_number": item.row_number,
                "order_no": item.order_no,
                "original_address_width": b2_text_width(item.original_address),
                "original_building_width": b2_text_width(item.original_building),
                "adjusted_address_width": b2_text_width(item.adjusted_address),
                "adjusted_building_width": b2_text_width(item.adjusted_building),
                "reasons": list(item.reasons),
                "auto_adjusted": item.auto_adjusted,
                "requires_review": item.requires_review,
            }
            for item in result.address_reviews
        ],
        "unmapped_item_codes": list(result.unmapped_item_codes),
        "warnings": list(result.warnings),
    }
    with result.audit_path.open("a", encoding="utf-8") as fp:
        fp.write(json.dumps(payload, ensure_ascii=False) + "\n")


def dataframe_preview(dataframe: pd.DataFrame, *, limit: int) -> tuple[dict[str, str], ...]:
    columns = [column for column in PREVIEW_COLUMNS if column in dataframe.columns]
    rows: list[dict[str, str]] = []
    for record in dataframe[columns].head(max(limit, 0)).to_dict(orient="records"):
        rows.append({key: normalize_cell(value) for key, value in record.items()})
    return tuple(rows)


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
