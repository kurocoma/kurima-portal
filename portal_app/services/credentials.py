from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from portal_app.services.excel_io import excel_read_snapshot


DEFAULT_CREDENTIAL_PATH = (
    Path.home() / "開発案件" / "日別売上集計データダウンロード" / "docs" / "ID・PW.xlsx"
)
# 認証情報Excel（ID・PW.xlsx）はネクストエンジン専用ではなく、楽天・Yahoo など
# 全サイトの行を持つ共有ファイル。歴史的経緯で NEXT_ENGINE_CREDENTIAL_PATH という
# 名前が先にあるが、用途が誤解されるため KURIMA_CREDENTIAL_PATH を正とし、
# 従来キーは後方互換として引き続き読む（両方あれば KURIMA_ 側が優先）。
CREDENTIAL_PATH_ENV = "KURIMA_CREDENTIAL_PATH"
LEGACY_CREDENTIAL_PATH_ENV = "NEXT_ENGINE_CREDENTIAL_PATH"


def credential_workbook_path() -> Path:
    """認証情報Excelのパスを解決する（KURIMA_ → 従来キー → 既定の順）。"""

    for env_name in (CREDENTIAL_PATH_ENV, LEGACY_CREDENTIAL_PATH_ENV):
        value = os.environ.get(env_name, "").strip()
        if value:
            return Path(value)
    return DEFAULT_CREDENTIAL_PATH


@dataclass(frozen=True)
class NextEngineCredential:
    login_id: str
    password: str


def load_next_engine_credential() -> NextEngineCredential:
    env_login_id = os.environ.get("NEXT_ENGINE_LOGIN_ID")
    env_password = os.environ.get("NEXT_ENGINE_PASSWORD")
    if env_login_id and env_password:
        return NextEngineCredential(login_id=env_login_id, password=env_password)

    return _load_from_workbook(credential_workbook_path())


def load_site_credential(site_label: str) -> tuple[str, str] | None:
    """認証情報Excel（ID・PW.xlsx）から任意サイト行の (ID, パスワード) を読む。

    A列がサイト名・B列がID・C列がパスワードという既存フォーマット
    （日別売上集計データダウンロードプロジェクトと共通）。
    見つからなければ None。値はログへ出さないこと。

    ネクストエンジン以外（楽天・Yahoo等）からも使うため、ファイルの場所は
    KURIMA_CREDENTIAL_PATH で指定する（従来の NEXT_ENGINE_CREDENTIAL_PATH も可）。
    """

    credential_path = credential_workbook_path()
    if not credential_path.exists():
        raise FileNotFoundError(f"認証情報ファイルが見つかりません: {credential_path}")

    # OneDrive等の共有フォルダ上でもロック衝突しないよう一時コピーを解析する。
    with excel_read_snapshot(credential_path) as snapshot:
        workbook = openpyxl.load_workbook(snapshot, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if str(row[0] or "").strip() != site_label:
                    continue
                login_id = str(row[1] or "").strip()
                password = str(row[2] or "").strip()
                if login_id and password:
                    return login_id, password
        finally:
            workbook.close()
    return None


def _load_from_workbook(path: Path) -> NextEngineCredential:
    if not path.exists():
        raise FileNotFoundError(
            "Next Engine の認証情報が見つかりません。"
            "NEXT_ENGINE_LOGIN_ID/NEXT_ENGINE_PASSWORD または "
            f"{CREDENTIAL_PATH_ENV}（従来キー: {LEGACY_CREDENTIAL_PATH_ENV}）"
            f"を設定してください: {path}"
        )

    # OneDrive等の共有フォルダ上でもロック衝突しないよう一時コピーを解析する。
    with excel_read_snapshot(path) as snapshot:
        workbook = openpyxl.load_workbook(snapshot, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                site_label = str(row[0] or "").strip()
                if site_label != "ネクストエンジン":
                    continue

                login_id = str(row[1] or "").strip()
                password = str(row[2] or "").strip()
                if not login_id or not password:
                    break
                return NextEngineCredential(login_id=login_id, password=password)
        finally:
            workbook.close()

    raise ValueError(f"認証情報ファイルにネクストエンジン行が見つかりません: {path}")
